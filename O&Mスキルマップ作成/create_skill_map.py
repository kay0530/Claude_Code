# -*- coding: utf-8 -*-
"""
O&Mグループ スキルマップ作成スクリプト
週報データに基づいたスキル評価とレーダーチャートを作成
"""

import openpyxl
from openpyxl.chart import RadarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# スキル評価項目
skill_categories = [
    "知識",
    "技術力",
    "リーダーシップ",
    "判断力",
    "応用力",
    "施工力",
    "点検力"
]

# 各メンバーのスキル評価（週報分析に基づく5点満点評価）
# 評価根拠：
# - 廣木徳文（54歳、創業メンバー、職人気質）: 電気工事の深い知識と経験、点検技術が高い
# - 太田賢宏（電気工事経験者）: 施工技術に優れ、リーダーシップも発揮、判断力も高い
# - 笹沼和宏（9月までパワまる営業）: 営業経験があり、施工は学習中、コミュニケーション能力高い
# - 田野隼人（使用前自己確認検査、自家消費現調多い）: 点検・調査能力に優れる
# - 和埜達人（25年4月新卒、高卒）: 学習意欲高く成長中、基礎スキルは発展途上
# - 山崎魁人（経験浅く入社、リーダー経験あり、26年2月リーダー辞退）: バランス良いスキル、リーダー経験から判断力あり
# - 淀川大智（元リーダー、業務委託、最も知識経験豊富）: 全項目で最高レベル

member_skills = {
    "廣木徳文": {
        "知識": 4.5,
        "技術力": 4.5,
        "リーダーシップ": 3.0,
        "判断力": 4.0,
        "応用力": 3.5,
        "施工力": 4.5,
        "点検力": 4.5
    },
    "太田賢宏": {
        "知識": 4.0,
        "技術力": 4.0,
        "リーダーシップ": 3.5,
        "判断力": 3.5,
        "応用力": 4.0,
        "施工力": 4.0,
        "点検力": 4.0
    },
    "笹沼和宏": {
        "知識": 2.5,
        "技術力": 2.5,
        "リーダーシップ": 2.5,
        "判断力": 2.5,
        "応用力": 3.0,
        "施工力": 2.5,
        "点検力": 2.5
    },
    "田野隼人": {
        "知識": 3.5,
        "技術力": 3.5,
        "リーダーシップ": 2.5,
        "判断力": 3.5,
        "応用力": 3.5,
        "施工力": 3.5,
        "点検力": 4.0
    },
    "和埜達人": {
        "知識": 2.0,
        "技術力": 2.0,
        "リーダーシップ": 1.5,
        "判断力": 2.0,
        "応用力": 2.5,
        "施工力": 2.0,
        "点検力": 2.0
    },
    "山崎魁人": {
        "知識": 4.0,
        "技術力": 4.0,
        "リーダーシップ": 3.5,
        "判断力": 4.0,
        "応用力": 4.0,
        "施工力": 4.0,
        "点検力": 4.0
    },
    "淀川大智": {
        "知識": 5.0,
        "技術力": 5.0,
        "リーダーシップ": 5.0,
        "判断力": 5.0,
        "応用力": 5.0,
        "施工力": 5.0,
        "点検力": 5.0
    }
}

# 評価根拠コメント
member_comments = {
    "廣木徳文": "創業メンバーとして豊富な電気工事経験を持つ。職人気質で技術力・点検力が高く、特に屋根上作業や電気配線の知識が深い。若手への技術指導も担当。",
    "太田賢宏": "電気工事経験者として入社。パワまる施工のリーダー的存在で、回路組換えや現調での判断力が高い。自己反省も多く、改善意識が強い。熱中症で休憩を取る場面も。",
    "笹沼和宏": "9月までパワまる営業を担当後、O&M業務に復帰。施工は学習中で作業スピードが課題。準備や部材確認のミスを反省し改善に取り組む姿勢あり。",
    "田野隼人": "使用前自己確認検査や自家消費現調を多数担当。点検業務に強みがあり、不良パネルの特定作業などで実績。地道な調査作業が得意。",
    "和埜達人": "2025年4月新卒入社（高卒）。学習意欲が高く、先輩から積極的に学ぶ姿勢。パワまる施工の基本作業を徐々に習得中。安全意識も高まりつつある。",
    "山崎魁人": "経験浅く入社後、リーダーを務めるまで成長。2026年2月にプレイヤースキル向上のためリーダー辞退。現調での最適解を考える力、後輩育成意識が高い。",
    "淀川大智": "元リーダー、現在は業務委託。O&Mグループで最も知識・経験が豊富。現場判断力が抜群で、工程見積もりや本質的な問題解決に長ける。全メンバーの指導者的存在。"
}

def create_skill_map_excel():
    """スキルマップExcelファイルを作成"""

    wb = openpyxl.Workbook()

    # === シート1: サマリー ===
    ws_summary = wb.active
    ws_summary.title = "スキルマップサマリー"

    # タイトル
    ws_summary['A1'] = "O&Mグループ スキルマップ"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary.merge_cells('A1:H1')

    ws_summary['A2'] = "作成日: 2026年2月"
    ws_summary['A3'] = "評価基準: 週報データ（2025年1月〜2026年1月）に基づく"

    # ヘッダー行
    headers = ["メンバー名"] + skill_categories + ["合計", "平均"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')

    # データ行
    row = 6
    for member, skills in member_skills.items():
        ws_summary.cell(row=row, column=1, value=member)
        col = 2
        total = 0
        for category in skill_categories:
            value = skills[category]
            ws_summary.cell(row=row, column=col, value=value)
            total += value
            col += 1
        ws_summary.cell(row=row, column=col, value=total)
        ws_summary.cell(row=row, column=col+1, value=round(total/len(skill_categories), 2))
        row += 1

    # 列幅調整
    ws_summary.column_dimensions['A'].width = 15
    for col in range(2, 11):
        ws_summary.column_dimensions[get_column_letter(col)].width = 12

    # === シート2: 評価コメント ===
    ws_comments = wb.create_sheet("評価コメント")
    ws_comments['A1'] = "メンバー別 評価コメント"
    ws_comments['A1'].font = Font(size=14, bold=True)

    row = 3
    for member, comment in member_comments.items():
        ws_comments.cell(row=row, column=1, value=member)
        ws_comments.cell(row=row, column=1).font = Font(bold=True)
        ws_comments.cell(row=row, column=2, value=comment)
        ws_comments.row_dimensions[row].height = 60
        row += 1

    ws_comments.column_dimensions['A'].width = 15
    ws_comments.column_dimensions['B'].width = 100

    # === シート3〜9: 各メンバーのレーダーチャート ===
    for member, skills in member_skills.items():
        ws = wb.create_sheet(member)

        # タイトル
        ws['A1'] = f"{member} スキルマップ"
        ws['A1'].font = Font(size=14, bold=True)

        # データテーブル
        ws['A3'] = "評価項目"
        ws['B3'] = "評価点"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)

        row = 4
        for category in skill_categories:
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=skills[category])
            row += 1

        # レーダーチャート作成
        chart = RadarChart()
        chart.type = "filled"
        chart.title = f"{member} スキルレーダー"
        chart.style = 10

        # データ参照
        labels = Reference(ws, min_col=1, min_row=4, max_row=4+len(skill_categories)-1)
        data = Reference(ws, min_col=2, min_row=3, max_row=3+len(skill_categories))

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        # チャートサイズ
        chart.width = 15
        chart.height = 12

        ws.add_chart(chart, "D3")

        # コメント追加
        ws['A15'] = "評価コメント:"
        ws['A15'].font = Font(bold=True)
        ws['A16'] = member_comments[member]
        ws.merge_cells('A16:J18')

        # 列幅調整
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 10

    # === シート10: 全員比較チャート用データ ===
    ws_all = wb.create_sheet("全員比較")

    ws_all['A1'] = "全員スキル比較"
    ws_all['A1'].font = Font(size=14, bold=True)

    # ヘッダー
    ws_all['A3'] = "評価項目"
    col = 2
    for member in member_skills.keys():
        ws_all.cell(row=3, column=col, value=member)
        ws_all.cell(row=3, column=col).font = Font(bold=True)
        col += 1

    # データ
    row = 4
    for category in skill_categories:
        ws_all.cell(row=row, column=1, value=category)
        col = 2
        for member, skills in member_skills.items():
            ws_all.cell(row=row, column=col, value=skills[category])
            col += 1
        row += 1

    # 全員比較レーダーチャート
    chart_all = RadarChart()
    chart_all.type = "marker"
    chart_all.title = "全員スキル比較"
    chart_all.style = 10

    labels_all = Reference(ws_all, min_col=1, min_row=4, max_row=4+len(skill_categories)-1)
    data_all = Reference(ws_all, min_col=2, min_row=3, max_col=2+len(member_skills)-1, max_row=3+len(skill_categories))

    chart_all.add_data(data_all, titles_from_data=True)
    chart_all.set_categories(labels_all)

    chart_all.width = 20
    chart_all.height = 15

    ws_all.add_chart(chart_all, "A12")

    # 列幅調整
    ws_all.column_dimensions['A'].width = 15
    for col in range(2, 9):
        ws_all.column_dimensions[get_column_letter(col)].width = 12

    # ファイル保存
    output_path = r"C:\Users\田中　圭亮\Desktop\Claude_Code_Demo\O&M_スキルマップ.xlsx"
    wb.save(output_path)
    print(f"スキルマップExcelファイルを作成しました: {output_path}")
    return output_path

if __name__ == "__main__":
    create_skill_map_excel()
