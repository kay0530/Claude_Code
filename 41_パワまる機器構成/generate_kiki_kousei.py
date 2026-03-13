# -*- coding: utf-8 -*-
"""
パワまる機器構成 テンプレート生成スクリプト

Phase 1: openpyxl で .xlsx を作成（シート・書式・データ・数式・ドロップダウン）
Phase 2: win32com で VBA コードを注入し、ボタンを追加して .xlsm として保存
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = SCRIPT_DIR
PVPCS_PATH = r"C:\Users\田中　圭亮\Desktop\PVPCSデータ.xlsx"
TEMP_XLSX = os.path.join(OUTPUT_DIR, "temp_template.xlsx")
OUTPUT_XLSM = os.path.join(OUTPUT_DIR, "パワまる機器構成_テンプレート.xlsm")
VBA_BAS_PATH = os.path.join(OUTPUT_DIR, "vba_code.bas")

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
FONT_DEFAULT = Font(name="Meiryo UI", size=11)
FONT_TITLE = Font(name="Meiryo UI", size=16, bold=True)
FONT_SECTION = Font(name="Meiryo UI", size=11, bold=True)
FONT_HEADER = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")

FILL_TITLE = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_SECTION_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FILL_SECTION_GRAY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
FILL_SECTION_RED = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
FILL_INPUT = PatternFill(start_color="FFFFDD", end_color="FFFFDD", fill_type="solid")
FILL_HEADER_ROW = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_FORMULA = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER_CONTINUOUS = Alignment(horizontal="centerContinuous", vertical="center")

THIN_SIDE = Side(style="thin")
BORDER_THIN = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------
def set_cell(ws, row, col, value, font=None, fill=None, alignment=None,
             border=None, merge_end_col=None, number_format=None):
    """Set a cell value with optional formatting and optional merge."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            mc = ws.cell(row=row, column=c)
            if fill:
                mc.fill = fill
            if border:
                mc.border = border
    return cell


# ===================================================================
# Phase 1 – openpyxl workbook creation
# ===================================================================

def create_workbook():
    """Create the complete workbook with all sheets."""
    wb = openpyxl.Workbook()
    create_input_sheet(wb)
    import_pv_data(wb)
    import_existing_pcs_data(wb)
    import_new_pcs_data(wb)
    create_cable_master(wb)
    create_breaker_master(wb)
    return wb


# ---------------------------------------------------------------------------
# Input sheet helpers
# ---------------------------------------------------------------------------

def _apply_input_cell_style(ws, row, col_start, col_end):
    """Apply input-cell style (thin border + yellow fill) to a range of cells."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_INPUT
        cell.border = BORDER_THIN
        cell.font = FONT_DEFAULT
        cell.alignment = ALIGN_CENTER


def _apply_formula_cell_style(ws, row, col_start, col_end):
    """Apply formula-cell style (thin border + green fill)."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_FORMULA
        cell.border = BORDER_THIN
        cell.font = FONT_DEFAULT
        cell.alignment = ALIGN_CENTER


def _set_header_row(ws, row, labels, max_col):
    """Set a header row with white bold text on blue background."""
    for i, label in enumerate(labels):
        col = i + 1
        set_cell(ws, row, col, label,
                 font=FONT_HEADER, fill=FILL_HEADER_ROW,
                 alignment=ALIGN_CENTER, border=BORDER_THIN)
    # Fill remaining columns if labels shorter than max_col
    for col in range(len(labels) + 1, max_col + 1):
        set_cell(ws, row, col, None,
                 font=FONT_HEADER, fill=FILL_HEADER_ROW,
                 alignment=ALIGN_CENTER, border=BORDER_THIN)


def _set_label_row(ws, row, label):
    """Set a label in column A with default style + border."""
    set_cell(ws, row, 1, label,
             font=FONT_DEFAULT, alignment=ALIGN_LEFT, border=BORDER_THIN)


def _apply_section_continuous(ws, row, end_col, font, fill):
    """Apply centerContinuous alignment across a section header row."""
    for c in range(2, end_col + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).alignment = ALIGN_CENTER_CONTINUOUS


# ---------------------------------------------------------------------------
# Input sheet
# ---------------------------------------------------------------------------

def create_input_sheet(wb):
    """Create the 入力 (Input) sheet."""
    ws = wb.active
    ws.title = "入力"

    # Column widths
    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, 14):  # B-M
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    # ── Row 1: Title ──
    set_cell(ws, 1, 1, "パワまる機器構成 入力シート",
             font=Font(name="Meiryo UI", size=16, bold=True, color="FFFFFF"),
             fill=FILL_TITLE,
             alignment=ALIGN_CENTER_CONTINUOUS)
    _apply_section_continuous(ws, 1, 13, FONT_TITLE, FILL_TITLE)
    ws.row_dimensions[1].height = 36

    # ── Row 3: 基本情報 ──
    set_cell(ws, 3, 1, "■ 基本情報", font=FONT_SECTION,
             fill=FILL_SECTION_BLUE, alignment=ALIGN_CENTER_CONTINUOUS)
    _apply_section_continuous(ws, 3, 13, FONT_SECTION, FILL_SECTION_BLUE)

    # Row 4-6: Basic info
    _set_label_row(ws, 4, "顧客名")
    _apply_input_cell_style(ws, 4, 2, 2)

    _set_label_row(ws, 5, "拠点名")
    _apply_input_cell_style(ws, 5, 2, 2)

    _set_label_row(ws, 6, "種別")
    _apply_input_cell_style(ws, 6, 2, 2)
    dv_type = DataValidation(type="list",
                             formula1='"パワまる三相,パワまる単相"',
                             allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_type)
    dv_type.add(ws["B6"])

    # ── Row 8: 既設側 ──
    set_cell(ws, 8, 1, "■ 既設側", font=FONT_SECTION,
             fill=FILL_SECTION_GRAY, alignment=ALIGN_CENTER_CONTINUOUS)
    _apply_section_continuous(ws, 8, 13, FONT_SECTION, FILL_SECTION_GRAY)

    # Row 9: PCS台数
    _set_label_row(ws, 9, "PCS台数")
    _apply_input_cell_style(ws, 9, 2, 2)
    dv_pcs_count = DataValidation(type="whole", operator="between",
                                  formula1="1", formula2="12",
                                  allow_blank=True, showErrorMessage=True)
    dv_pcs_count.error = "1〜12の整数を入力してください"
    dv_pcs_count.errorTitle = "入力エラー"
    ws.add_data_validation(dv_pcs_count)
    dv_pcs_count.add(ws["B9"])

    # Row 10: パネル枚数(合計)
    _set_label_row(ws, 10, "パネル枚数(合計)")
    _apply_input_cell_style(ws, 10, 2, 2)
    dv_panel_count = DataValidation(type="whole", operator="greaterThan",
                                    formula1="0",
                                    allow_blank=True, showErrorMessage=True)
    dv_panel_count.error = "正の整数を入力してください"
    dv_panel_count.errorTitle = "入力エラー"
    ws.add_data_validation(dv_panel_count)
    dv_panel_count.add(ws["B10"])

    # ── Row 12: Header row (existing PCS table) ──
    existing_headers = ["項目"] + [f"PCS{i}" for i in range(1, 13)]
    _set_header_row(ws, 12, existing_headers, 13)

    # Rows 13-28: Existing PCS detail rows (v2 row numbers)
    existing_labels = {
        13: "パネル型番",
        14: "直列数",
        15: "集電数",
        16: "回路数",
        17: "直列数2",
        18: "集電数2",
        19: "回路数2",
        # 20: 合計枚数 (formula row, handled separately)
        21: "※想定",
        22: "PCS型番",
        23: "PCS方式",
        24: "PCS容量(kW)",
        25: "ケーブル種(1段)",
        26: "ブレーカ種",
        27: "ブレーカ容量",
        28: "感度電流",
    }
    for row_num, label in existing_labels.items():
        _set_label_row(ws, row_num, label)

    # Input cells for rows 13-19, 21-22, 25-28 (B-M)
    for row_num in [13, 14, 15, 16, 17, 18, 19, 21, 22, 25, 26, 27, 28]:
        _apply_input_cell_style(ws, row_num, 2, 13)

    # Row 20: 合計枚数 (formula row)
    _set_label_row(ws, 20, "合計枚数")
    _apply_formula_cell_style(ws, 20, 2, 13)
    for col in range(2, 14):
        cl = get_column_letter(col)
        ws.cell(row=20, column=col).value = f'=IF({cl}14="","",{cl}14*{cl}16+IF({cl}17="",0,{cl}17*{cl}19))'

    # Formula cells for rows 23-24 (green fill)
    for row_num in [23, 24]:
        _apply_formula_cell_style(ws, row_num, 2, 13)

    # Row 13: Panel model dropdown (PV!A2:A1400)
    dv_panel = DataValidation(type="list",
                              formula1="=PV!$A$2:$A$1400",
                              allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_panel)
    dv_panel.add("B13:M13")

    # Row 21: ※想定 dropdown
    dv_estimate = DataValidation(type="list",
                                 formula1='"○,"',
                                 allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_estimate)
    dv_estimate.add("B21:M21")

    # Row 22: PCS型番 dropdown (既設PCS!A2:A81)
    dv_existing_pcs = DataValidation(type="list",
                                     formula1="=既設PCS!$A$2:$A$81",
                                     allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_existing_pcs)
    dv_existing_pcs.add("B22:M22")

    # Row 23: VLOOKUP for PCS方式 (references PCS型番 at row 22)
    for col in range(2, 14):
        cl = get_column_letter(col)
        ws.cell(row=23, column=col).value = \
            f'=IFERROR(VLOOKUP({cl}22,既設PCS!A:B,2,FALSE),"")'

    # Row 24: VLOOKUP for PCS容量 (references PCS型番 at row 22)
    for col in range(2, 14):
        cl = get_column_letter(col)
        ws.cell(row=24, column=col).value = \
            f'=IFERROR(VLOOKUP({cl}22,既設PCS!A:D,4,FALSE),"")'

    # Row 25: Cable dropdown
    dv_cable = DataValidation(type="list",
                              formula1="=ケーブルマスタ!$A$2:$A$20",
                              allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_cable)
    dv_cable.add("B25:M25")

    # Row 26: Breaker type dropdown
    dv_brk_type = DataValidation(type="list",
                                 formula1="=ブレーカマスタ!$A$2:$A$3",
                                 allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_brk_type)
    dv_brk_type.add("B26:M26")

    # Row 27: Breaker capacity dropdown
    dv_brk_cap = DataValidation(type="list",
                                formula1="=ブレーカマスタ!$C$2:$C$8",
                                allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_brk_cap)
    dv_brk_cap.add("B27:M27")

    # Row 28: Sensitivity current dropdown
    dv_sensitivity = DataValidation(type="list",
                                    formula1="=ブレーカマスタ!$G$2:$G$5",
                                    allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_sensitivity)
    dv_sensitivity.add("B28:M28")

    # ── Row 30: 既設 集電盤 ──
    set_cell(ws, 30, 1, "■ 既設 集電盤", font=FONT_SECTION,
             fill=FILL_SECTION_GRAY, alignment=ALIGN_CENTER_CONTINUOUS)
    _apply_section_continuous(ws, 30, 13, FONT_SECTION, FILL_SECTION_GRAY)

    agg_labels = {
        31: "ケーブル種(2段)",
        32: "ブレーカ種(集電)",
        33: "ブレーカ容量(集電)",
        34: "感度電流(集電)",
        35: "幹線ケーブル",
    }
    for row_num, label in agg_labels.items():
        _set_label_row(ws, row_num, label)
        _apply_input_cell_style(ws, row_num, 2, 2)

    # Dropdowns for existing aggregate panel
    dv_cable2 = DataValidation(type="list",
                               formula1="=ケーブルマスタ!$A$2:$A$20",
                               allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_cable2)
    dv_cable2.add(ws["B31"])

    dv_brk_type_agg = DataValidation(type="list",
                                     formula1="=ブレーカマスタ!$A$2:$A$3",
                                     allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_brk_type_agg)
    dv_brk_type_agg.add(ws["B32"])

    dv_brk_cap_agg = DataValidation(type="list",
                                    formula1="=ブレーカマスタ!$E$2:$E$10",
                                    allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_brk_cap_agg)
    dv_brk_cap_agg.add(ws["B33"])

    dv_sens_agg = DataValidation(type="list",
                                 formula1="=ブレーカマスタ!$G$2:$G$5",
                                 allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_sens_agg)
    dv_sens_agg.add(ws["B34"])

    dv_trunk_cable = DataValidation(type="list",
                                    formula1="=ケーブルマスタ!$C$2:$C$20",
                                    allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_trunk_cable)
    dv_trunk_cable.add(ws["B35"])

    # ── Row 37: 新設側 ──
    set_cell(ws, 37, 1, "■ 新設側", font=FONT_SECTION,
             fill=FILL_SECTION_RED, alignment=ALIGN_CENTER_CONTINUOUS)
    _apply_section_continuous(ws, 37, 13, FONT_SECTION, FILL_SECTION_RED)

    # Row 38: PCS台数
    _set_label_row(ws, 38, "PCS台数")
    _apply_input_cell_style(ws, 38, 2, 2)
    dv_new_pcs_count = DataValidation(type="whole", operator="between",
                                      formula1="1", formula2="8",
                                      allow_blank=True, showErrorMessage=True)
    dv_new_pcs_count.error = "1〜8の整数を入力してください"
    dv_new_pcs_count.errorTitle = "入力エラー"
    ws.add_data_validation(dv_new_pcs_count)
    dv_new_pcs_count.add(ws["B38"])

    # Row 39: パネル枚数(合計) — formula reference to B10
    _set_label_row(ws, 39, "パネル枚数(合計)")
    ws.cell(row=39, column=2).value = "=B10"
    ws.cell(row=39, column=2).fill = FILL_FORMULA
    ws.cell(row=39, column=2).border = BORDER_THIN
    ws.cell(row=39, column=2).font = FONT_DEFAULT
    ws.cell(row=39, column=2).alignment = ALIGN_CENTER

    # ── Row 41: Header row (new PCS table) ──
    new_headers = ["項目"] + [f"PCS{i}" for i in range(1, 9)]
    _set_header_row(ws, 41, new_headers, 9)

    # Rows 42-56: New PCS detail rows (v2 row numbers)
    new_labels = {
        42: "パネル型番",
        43: "直列数",
        44: "集電数",
        45: "回路数",
        46: "直列数2",
        47: "集電数2",
        48: "回路数2",
        # 49: 合計枚数 (formula row, handled separately)
        50: "PCS型番",
        51: "PCS方式",
        52: "PCS容量(kW)",
        53: "ケーブル種(1段)",
        54: "ブレーカ種",
        55: "ブレーカ容量",
        56: "感度電流",
    }
    for row_num, label in new_labels.items():
        _set_label_row(ws, row_num, label)

    # Input cells for rows 42-48, 50, 53-56 (B-I)
    for row_num in [42, 43, 44, 45, 46, 47, 48, 50, 53, 54, 55, 56]:
        _apply_input_cell_style(ws, row_num, 2, 9)

    # Row 49: 合計枚数 (formula row)
    _set_label_row(ws, 49, "合計枚数")
    _apply_formula_cell_style(ws, 49, 2, 9)
    for col in range(2, 10):
        cl = get_column_letter(col)
        ws.cell(row=49, column=col).value = f'=IF({cl}43="","",{cl}43*{cl}45+IF({cl}46="",0,{cl}46*{cl}48))'

    # Formula cells for rows 51-52
    for row_num in [51, 52]:
        _apply_formula_cell_style(ws, row_num, 2, 9)

    # Row 42: Panel model dropdown
    dv_new_panel = DataValidation(type="list",
                                  formula1="=PV!$A$2:$A$1400",
                                  allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_panel)
    dv_new_panel.add("B42:I42")

    # Row 50: PCS model dropdown (変更後PCS!A2:A12)
    dv_new_pcs = DataValidation(type="list",
                                formula1="=変更後PCS!$A$2:$A$12",
                                allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_pcs)
    dv_new_pcs.add("B50:I50")

    # Row 51: VLOOKUP PCS方式 (references PCS型番 at row 50)
    for col in range(2, 10):
        cl = get_column_letter(col)
        ws.cell(row=51, column=col).value = \
            f'=IFERROR(VLOOKUP({cl}50,変更後PCS!A:B,2,FALSE),"")'

    # Row 52: VLOOKUP PCS容量 (references PCS型番 at row 50)
    for col in range(2, 10):
        cl = get_column_letter(col)
        ws.cell(row=52, column=col).value = \
            f'=IFERROR(VLOOKUP({cl}50,変更後PCS!A:D,4,FALSE),"")'

    # Row 53: Cable dropdown
    dv_new_cable = DataValidation(type="list",
                                  formula1="=ケーブルマスタ!$A$2:$A$20",
                                  allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_cable)
    dv_new_cable.add("B53:I53")

    # Row 54: Breaker type dropdown
    dv_new_brk_type = DataValidation(type="list",
                                     formula1="=ブレーカマスタ!$A$2:$A$3",
                                     allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_brk_type)
    dv_new_brk_type.add("B54:I54")

    # Row 55: Breaker capacity dropdown
    dv_new_brk_cap = DataValidation(type="list",
                                    formula1="=ブレーカマスタ!$C$2:$C$8",
                                    allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_brk_cap)
    dv_new_brk_cap.add("B55:I55")

    # Row 56: Sensitivity dropdown
    dv_new_sens = DataValidation(type="list",
                                 formula1="=ブレーカマスタ!$G$2:$G$5",
                                 allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_sens)
    dv_new_sens.add("B56:I56")

    # ── Row 58: 新設 集電盤 ──
    set_cell(ws, 58, 1, "■ 新設 集電盤", font=FONT_SECTION,
             fill=FILL_SECTION_RED, alignment=ALIGN_CENTER_CONTINUOUS)
    _apply_section_continuous(ws, 58, 9, FONT_SECTION, FILL_SECTION_RED)

    new_agg_labels = {
        59: "ケーブル種(2段)",
        60: "ブレーカ種(集電)",
        61: "ブレーカ容量(集電)",
        62: "感度電流(集電)",
        63: "幹線ケーブル",
    }
    for row_num, label in new_agg_labels.items():
        _set_label_row(ws, row_num, label)
        _apply_input_cell_style(ws, row_num, 2, 2)

    # Dropdowns for new aggregate panel
    dv_new_cable2 = DataValidation(type="list",
                                   formula1="=ケーブルマスタ!$A$2:$A$20",
                                   allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_cable2)
    dv_new_cable2.add(ws["B59"])

    dv_new_brk_type_agg = DataValidation(type="list",
                                         formula1="=ブレーカマスタ!$A$2:$A$3",
                                         allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_brk_type_agg)
    dv_new_brk_type_agg.add(ws["B60"])

    dv_new_brk_cap_agg = DataValidation(type="list",
                                        formula1="=ブレーカマスタ!$E$2:$E$10",
                                        allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_brk_cap_agg)
    dv_new_brk_cap_agg.add(ws["B61"])

    dv_new_sens_agg = DataValidation(type="list",
                                     formula1="=ブレーカマスタ!$G$2:$G$5",
                                     allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_sens_agg)
    dv_new_sens_agg.add(ws["B62"])

    dv_new_trunk = DataValidation(type="list",
                                  formula1="=ケーブルマスタ!$C$2:$C$20",
                                  allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_new_trunk)
    dv_new_trunk.add(ws["B63"])

    # ── Row 65: 備考 ──
    set_cell(ws, 65, 1, "■ 備考", font=FONT_SECTION,
             fill=FILL_SECTION_BLUE, alignment=ALIGN_CENTER_CONTINUOUS)
    _apply_section_continuous(ws, 65, 13, FONT_SECTION, FILL_SECTION_BLUE)

    remark_labels = {
        66: "回路組換",
        67: "直流線延長",
        68: "ケーブル張替え",
        69: "流用箇所メモ",
    }
    for row_num, label in remark_labels.items():
        _set_label_row(ws, row_num, label)
        _apply_input_cell_style(ws, row_num, 2, 2)

    # Row 66-68: ○ or blank dropdown
    dv_yn = DataValidation(type="list",
                           formula1='"○,"',
                           allow_blank=True, showErrorMessage=False)
    ws.add_data_validation(dv_yn)
    dv_yn.add(ws["B66"])
    dv_yn.add(ws["B67"])
    dv_yn.add(ws["B68"])

    # Row 69: Free text - just input cell, no dropdown

    # Freeze panes at row 12 (header row)
    ws.freeze_panes = "A12"

    # Sheet tab color
    ws.sheet_properties.tabColor = "4472C4"

    print("  [OK] 入力 sheet created")


# ---------------------------------------------------------------------------
# Import PV data
# ---------------------------------------------------------------------------

def import_pv_data(wb):
    """Import PV panel data from source file."""
    src = openpyxl.load_workbook(PVPCS_PATH, data_only=True)
    src_ws = src["PV"]
    ws = wb.create_sheet("PV")

    max_row = src_ws.max_row
    max_col = src_ws.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val = src_ws.cell(row=row, column=col).value
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = FONT_DEFAULT
            if row == 1:
                cell.font = Font(name="Meiryo UI", size=11, bold=True)
                cell.fill = FILL_HEADER_ROW
                cell.font = FONT_HEADER
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN

    # Auto-fit column widths (approximate)
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    ws.sheet_properties.tabColor = "70AD47"
    src.close()
    print(f"  [OK] PV sheet imported ({max_row - 1} rows)")


# ---------------------------------------------------------------------------
# Import existing PCS data
# ---------------------------------------------------------------------------

def import_existing_pcs_data(wb):
    """Import existing PCS data from source file."""
    src = openpyxl.load_workbook(PVPCS_PATH, data_only=True)
    src_ws = src["既設PCS"]
    ws = wb.create_sheet("既設PCS")

    max_row = src_ws.max_row
    max_col = src_ws.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val = src_ws.cell(row=row, column=col).value
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = FONT_DEFAULT
            if row == 1:
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER_ROW
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN

    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.sheet_properties.tabColor = "FFC000"
    src.close()
    print(f"  [OK] 既設PCS sheet imported ({max_row - 1} rows)")


# ---------------------------------------------------------------------------
# Import new PCS data
# ---------------------------------------------------------------------------

def import_new_pcs_data(wb):
    """Import post-change PCS data from source file."""
    src = openpyxl.load_workbook(PVPCS_PATH, data_only=True)
    src_ws = src["変更後PCS"]
    ws = wb.create_sheet("変更後PCS")

    max_row = src_ws.max_row
    max_col = src_ws.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val = src_ws.cell(row=row, column=col).value
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = FONT_DEFAULT
            if row == 1:
                cell.font = FONT_HEADER
                cell.fill = FILL_HEADER_ROW
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_THIN

    ws.column_dimensions["A"].width = 32
    for col_idx in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    ws.sheet_properties.tabColor = "FF6600"
    src.close()
    print(f"  [OK] 変更後PCS sheet imported ({max_row - 1} rows)")


# ---------------------------------------------------------------------------
# Cable master
# ---------------------------------------------------------------------------

def create_cable_master(wb):
    """Create the ケーブルマスタ sheet."""
    ws = wb.create_sheet("ケーブルマスタ")

    # Column A: PCS～ブレーカ間
    set_cell(ws, 1, 1, "PCS〜ブレーカ間",
             font=FONT_HEADER, fill=FILL_HEADER_ROW,
             alignment=ALIGN_CENTER, border=BORDER_THIN)
    cable_a = [
        "CV 8sq", "CV 14sq", "CV 22sq", "CV 38sq",
        "CV 60sq", "CV 100sq", "CV 150sq", "CV 200sq",
        "CV 250sq", "CV 325sq",
        "CVT 8sq", "CVT 14sq", "CVT 22sq", "CVT 38sq",
        "CVT 60sq", "CVT 100sq", "CVT 150sq", "CVT 200sq",
        "CVT 250sq",
    ]
    for i, val in enumerate(cable_a):
        set_cell(ws, i + 2, 1, val, font=FONT_DEFAULT, border=BORDER_THIN)

    # Column C: 幹線ケーブル
    set_cell(ws, 1, 3, "幹線ケーブル",
             font=FONT_HEADER, fill=FILL_HEADER_ROW,
             alignment=ALIGN_CENTER, border=BORDER_THIN)
    cable_c = [
        "CV 8sq", "CV 14sq", "CV 22sq", "CV 38sq", "CV 60sq",
        "CV 100sq", "CV 150sq", "CV 200sq",
        "CV 250sq", "CV 325sq",
        "CVT 8sq", "CVT 14sq", "CVT 22sq", "CVT 38sq", "CVT 60sq",
        "CVT 100sq", "CVT 150sq", "CVT 200sq", "CVT 250sq",
    ]
    for i, val in enumerate(cable_c):
        set_cell(ws, i + 2, 3, val, font=FONT_DEFAULT, border=BORDER_THIN)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.sheet_properties.tabColor = "A5A5A5"
    print("  [OK] ケーブルマスタ sheet created")


# ---------------------------------------------------------------------------
# Breaker master
# ---------------------------------------------------------------------------

def create_breaker_master(wb):
    """Create the ブレーカマスタ sheet."""
    ws = wb.create_sheet("ブレーカマスタ")

    # Column A: ブレーカ種別
    set_cell(ws, 1, 1, "ブレーカ種別",
             font=FONT_HEADER, fill=FILL_HEADER_ROW,
             alignment=ALIGN_CENTER, border=BORDER_THIN)
    set_cell(ws, 2, 1, "MCCB", font=FONT_DEFAULT, border=BORDER_THIN)
    set_cell(ws, 3, 1, "ELCB", font=FONT_DEFAULT, border=BORDER_THIN)

    # Column C: ブレーカ容量(1段)
    set_cell(ws, 1, 3, "ブレーカ容量(1段)",
             font=FONT_HEADER, fill=FILL_HEADER_ROW,
             alignment=ALIGN_CENTER, border=BORDER_THIN)
    cap_1 = [30, 40, 50, 75, 100, 150, 200]
    for i, val in enumerate(cap_1):
        set_cell(ws, i + 2, 3, val, font=FONT_DEFAULT, border=BORDER_THIN)

    # Column E: ブレーカ容量(集電盤)
    set_cell(ws, 1, 5, "ブレーカ容量(集電盤)",
             font=FONT_HEADER, fill=FILL_HEADER_ROW,
             alignment=ALIGN_CENTER, border=BORDER_THIN)
    cap_agg = [30, 40, 50, 75, 100, 150, 200, 250, 300]
    for i, val in enumerate(cap_agg):
        set_cell(ws, i + 2, 5, val, font=FONT_DEFAULT, border=BORDER_THIN)

    # Column G: 感度電流
    set_cell(ws, 1, 7, "感度電流",
             font=FONT_HEADER, fill=FILL_HEADER_ROW,
             alignment=ALIGN_CENTER, border=BORDER_THIN)
    sensitivity = [30, 100, 200, 500]
    for i, val in enumerate(sensitivity):
        set_cell(ws, i + 2, 7, val, font=FONT_DEFAULT, border=BORDER_THIN)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["G"].width = 14
    ws.sheet_properties.tabColor = "A5A5A5"
    print("  [OK] ブレーカマスタ sheet created")


# ===================================================================
# Phase 2 – win32com VBA injection
# ===================================================================

def inject_vba_and_save(xlsx_path, xlsm_path, vba_bas_path):
    """Open the .xlsx via Excel COM, inject VBA, add buttons, save as .xlsm."""
    import win32com.client

    if not os.path.exists(vba_bas_path):
        print(f"  [WARN] VBA file not found: {vba_bas_path}")
        print("  Skipping Phase 2 (VBA injection). Rename .xlsx manually if needed.")
        return

    print("  Starting Excel COM...")
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    try:
        abs_xlsx = os.path.abspath(xlsx_path)
        abs_xlsm = os.path.abspath(xlsm_path)

        wb = excel.Workbooks.Open(abs_xlsx)

        # Read VBA source code, stripping Attribute lines (file metadata
        # that is invalid inside AddFromString)
        with open(vba_bas_path, "r", encoding="utf-8") as f:
            vba_lines = f.readlines()
        vba_code = "".join(
            line for line in vba_lines
            if not line.strip().startswith("Attribute ")
        )

        # Add standard module
        vb_module = wb.VBProject.VBComponents.Add(1)  # vbext_ct_StdModule
        vb_module.Name = "ModKikiKousei"
        vb_module.CodeModule.AddFromString(vba_code)

        # Add UserForm for estimate side selection
        vb_form = wb.VBProject.VBComponents.Add(3)  # vbext_ct_MSForm
        vb_form.Name = "frmEstimateSide"

        # Set form properties
        vb_form.Properties("Caption").Value = "ストリング構成推定"
        vb_form.Properties("Width").Value = 300
        vb_form.Properties("Height").Value = 150

        # Add buttons via designer
        designer = vb_form.Designer

        btn_existing = designer.Controls.Add("Forms.CommandButton.1")
        btn_existing.Name = "btnExisting"
        btn_existing.Caption = "既設側"
        btn_existing.Left = 20
        btn_existing.Top = 40
        btn_existing.Width = 75
        btn_existing.Height = 30

        btn_new = designer.Controls.Add("Forms.CommandButton.1")
        btn_new.Name = "btnNew"
        btn_new.Caption = "新設側"
        btn_new.Left = 110
        btn_new.Top = 40
        btn_new.Width = 75
        btn_new.Height = 30

        btn_cancel = designer.Controls.Add("Forms.CommandButton.1")
        btn_cancel.Name = "btnCancel"
        btn_cancel.Caption = "キャンセル"
        btn_cancel.Left = 200
        btn_cancel.Top = 40
        btn_cancel.Width = 75
        btn_cancel.Height = 30

        # Add label
        lbl = designer.Controls.Add("Forms.Label.1")
        lbl.Name = "lblPrompt"
        lbl.Caption = "推定する側を選択してください:"
        lbl.Left = 20
        lbl.Top = 15
        lbl.Width = 260
        lbl.Height = 20

        # Add form code
        form_code = (
            'Public gEstimateSide As String\n'
            '\n'
            'Private Sub UserForm_Initialize()\n'
            '    gEstimateSide = "cancel"\n'
            'End Sub\n'
            '\n'
            'Private Sub btnExisting_Click()\n'
            '    gEstimateSide = "existing"\n'
            '    Me.Hide\n'
            'End Sub\n'
            '\n'
            'Private Sub btnNew_Click()\n'
            '    gEstimateSide = "new"\n'
            '    Me.Hide\n'
            'End Sub\n'
            '\n'
            'Private Sub btnCancel_Click()\n'
            '    gEstimateSide = "cancel"\n'
            '    Me.Hide\n'
            'End Sub\n'
        )
        vb_form.CodeModule.AddFromString(form_code)

        # Add Worksheet_Change event code to the input sheet's code module
        ws_input = wb.Sheets("入力")
        ws_code = wb.VBProject.VBComponents(ws_input.CodeName).CodeModule

        worksheet_change_code = (
            "Private Sub Worksheet_Change(ByVal Target As Range)\n"
            "    ' Update new panel model dropdown based on existing panel selections\n"
            "    If Not Intersect(Target, Me.Range(\"B13:M13\")) Is Nothing Then\n"
            "        Application.EnableEvents = False\n"
            "        On Error GoTo CleanUp\n"
            "        \n"
            "        ' Collect unique panel models from existing side\n"
            "        Dim dict As Object\n"
            "        Set dict = CreateObject(\"Scripting.Dictionary\")\n"
            "        Dim c As Long\n"
            "        For c = 2 To 13  ' B to M\n"
            "            Dim v As String\n"
            "            v = Trim(CStr(Me.Cells(13, c).Value))\n"
            "            If v <> \"\" And Not dict.Exists(v) Then\n"
            "                dict.Add v, True\n"
            "            End If\n"
            "        Next c\n"
            "        \n"
            "        ' Build validation list\n"
            "        Dim validList As String\n"
            "        If dict.Count > 0 Then\n"
            "            Dim keys As Variant\n"
            "            keys = dict.keys\n"
            "            Dim k As Long\n"
            "            For k = 0 To UBound(keys)\n"
            "                If k > 0 Then validList = validList & \",\"\n"
            "                validList = validList & keys(k)\n"
            "            Next k\n"
            "        End If\n"
            "        \n"
            "        ' Update validation on new panel model cells (row 42, B-I)\n"
            "        Dim rng As Range\n"
            "        Set rng = Me.Range(\"B42:I42\")\n"
            "        rng.Validation.Delete\n"
            "        If validList <> \"\" Then\n"
            "            rng.Validation.Add Type:=xlValidateList, AlertStyle:=xlValidAlertStop, _\n"
            "                Formula1:=validList\n"
            "            rng.Validation.IgnoreBlank = True\n"
            "            rng.Validation.InCellDropdown = True\n"
            "            rng.Validation.ShowError = False\n"
            "        End If\n"
            "        \n"
            "CleanUp:\n"
            "        Application.EnableEvents = True\n"
            "    End If\n"
            "End Sub\n"
        )
        ws_code.AddFromString(worksheet_change_code)

        # Add "作成" button on input sheet (row 71 area)
        left = ws_input.Range("B71").Left
        top = ws_input.Range("B71").Top
        btn = ws_input.Buttons().Add(left, top, 200, 40)
        btn.OnAction = "GenerateKikiKousei"
        btn.Caption = "\u4f5c\u6210"  # 作成
        btn.Characters.Font.Size = 14
        btn.Characters.Font.Bold = True

        # Add "推定" button next to "作成" button
        left2 = ws_input.Range("B71").Left + 220
        top2 = ws_input.Range("B71").Top
        btn2 = ws_input.Buttons().Add(left2, top2, 200, 40)
        btn2.OnAction = "EstimateStringConfig"
        btn2.Caption = "\u63a8\u5b9a"  # 推定
        btn2.Characters.Font.Size = 14
        btn2.Characters.Font.Bold = True

        # Save as .xlsm (macro-enabled: FileFormat=52)
        wb.SaveAs(abs_xlsm, FileFormat=52)
        wb.Close()
        print(f"  [OK] Saved .xlsm: {abs_xlsm}")

    except Exception as e:
        print(f"  [ERROR] Phase 2 failed: {e}")
        raise
    finally:
        try:
            excel.Quit()
        except Exception:
            pass


# ===================================================================
# Main
# ===================================================================

def main():
    print("=" * 60)
    print("パワまる機器構成 テンプレート生成")
    print("=" * 60)

    # Validate source file
    if not os.path.exists(PVPCS_PATH):
        print(f"[ERROR] Source file not found: {PVPCS_PATH}")
        sys.exit(1)

    # Phase 1: Create workbook with openpyxl
    print("\n--- Phase 1: Creating workbook (openpyxl) ---")
    wb = create_workbook()
    wb.save(TEMP_XLSX)
    print(f"\n  Phase 1 complete: {TEMP_XLSX}")

    # Phase 2: Inject VBA and save as .xlsm
    print("\n--- Phase 2: Injecting VBA (win32com) ---")
    try:
        inject_vba_and_save(TEMP_XLSX, OUTPUT_XLSM, VBA_BAS_PATH)
    except ImportError:
        print("  [WARN] win32com not available. Saving as .xlsx only.")
        # Rename to .xlsm anyway (won't have macros)
        import shutil
        fallback_xlsx = OUTPUT_XLSM.replace(".xlsm", ".xlsx")
        shutil.copy2(TEMP_XLSX, fallback_xlsx)
        print(f"  Saved as: {fallback_xlsx}")
    except Exception as e:
        print(f"  [WARN] Phase 2 failed: {e}")
        print("  The .xlsx file is available for manual processing.")

    # Cleanup temp file
    if os.path.exists(TEMP_XLSX):
        import time
        for attempt in range(5):
            try:
                os.remove(TEMP_XLSX)
                print("\n  Temp file cleaned up.")
                break
            except PermissionError:
                time.sleep(1)
        else:
            print(f"\n  [WARN] Could not remove temp file: {TEMP_XLSX}")

    print("\n" + "=" * 60)
    print("Done!")
    print("=" * 60)


if __name__ == "__main__":
    main()
