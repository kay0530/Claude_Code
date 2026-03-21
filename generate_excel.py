# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

output_path = os.path.join(r"C:\Users\田中　圭亮\Desktop\Claude_Code_Demo", "商品企画アイデア100_5エージェント昇華版.xlsx")

wb = openpyxl.Workbook()

# === Styles ===
title_font = Font(name="Meiryo UI", size=16, bold=True, color="FFFFFF")
title_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
h2_font = Font(name="Meiryo UI", size=13, bold=True, color="1F4E79")
h3_font = Font(name="Meiryo UI", size=11, bold=True, color="2E75B6")
header_font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
body_font = Font(name="Meiryo UI", size=10)
body_font_s = Font(name="Meiryo UI", size=9)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
theme_fills = {
    "A": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "B": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "C": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "D": PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
    "E": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "F": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
}

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        for c in range(col, merge_end_col + 1):
            mc = ws.cell(row=row, column=c)
            if fill:
                mc.fill = fill
            if border:
                mc.border = border
    return cell

# ====================================================================
# Sheet 1: Executive Summary
# ====================================================================
ws1 = wb.active
ws1.title = "1_エグゼクティブサマリー"
ws1.sheet_properties.tabColor = "1F4E79"

for col_idx in range(1, 8):
    ws1.column_dimensions[get_column_letter(col_idx)].width = 25

r = 1
set_cell(ws1, r, 1, "商品企画アイデア100 - 5エージェント昇華レポート", title_font, title_fill, center_align, merge_end_col=7)
r += 2

set_cell(ws1, r, 1, "【分析概要】", h2_font)
r += 1
set_cell(ws1, r, 1, "5つの専門AIエージェントが、太陽光発電O&Mサービス（エナジーHUB / パワまる）の商品企画アイデア100本を、", body_font, alignment=wrap_align)
r += 1
set_cell(ws1, r, 1, "「顧客が本当に困っていること」を起点に再構築しました。", body_font, alignment=wrap_align)
r += 2

# Structural Problems
set_cell(ws1, r, 1, "■ 発見された3つの構造的バイアス", h2_font)
r += 1
biases = [
    ["バイアス1", "可視化ゴール症候群", "「見える化」自体がゴールになり、顧客の行動変容につながらない"],
    ["バイアス2", "技術主語の罠", "「AIが検知」「IoTで連携」など技術が主語で、顧客の課題解決が後回し"],
    ["バイアス3", "営業目的の隠蔽", "顧客メリットに見せかけて、実はアップセル・クロスセル目的が透けている"],
]
headers_b = ["No", "バイアス名", "内容"]
for ci, h in enumerate(headers_b, 1):
    set_cell(ws1, r, ci, h, header_font, header_fill, center_align, thin_border)
r += 1
for bias in biases:
    for ci, v in enumerate(bias, 1):
        set_cell(ws1, r, ci, v, body_font, alignment=wrap_align, border=thin_border)
    r += 1
r += 1

# 5 Pains
set_cell(ws1, r, 1, "■ 顧客が本当に抱える5つの潜在ペイン", h2_font)
r += 1
pains = [
    ["P1", "発電量低下の不安", "正常なのか異常なのかわからず、ズルズル損失を放置してしまう", "kWh損失の金額可視化"],
    ["P2", "業者依存の不透明さ", "メンテ業者の言い値が適正か判断できず、不信感を抱えている", "相場比較・作業証跡"],
    ["P3", "制度変更の恐怖", "FIT終了・法改正・税制変更への対処法がわからず不安", "影響シミュレーション"],
    ["P4", "出口戦略の欠如", "設備の売却・延命・撤去のどれが得か判断基準がない", "経済比較ツール"],
    ["P5", "経営指標との断絶", "発電データと経営数字（P/L, CF）がつながらず、投資判断ができない", "経営ダッシュボード連携"],
]
headers_p = ["No", "ペイン名", "顧客の本音", "解決の方向性"]
for ci, h in enumerate(headers_p, 1):
    set_cell(ws1, r, ci, h, header_font, header_fill, center_align, thin_border)
r += 1
for pain in pains:
    for ci, v in enumerate(pain, 1):
        set_cell(ws1, r, ci, v, body_font, alignment=wrap_align, border=thin_border)
    r += 1
r += 1

# 6 Principles
set_cell(ws1, r, 1, "■ 昇華のための6つの設計原則", h2_font)
r += 1
principles = [
    ["原則1", "金額で語れ", "kWhやPR値ではなく「月○万円の損失」で伝える"],
    ["原則2", "比較させろ", "他の発電所・業界平均との比較で危機感を持たせる"],
    ["原則3", "次の行動を示せ", "アラートだけでなく「今すぐやること」を提示する"],
    ["原則4", "お金の流れで見せろ", "発電データと経営（P/L, CF, 税）を直結させる"],
    ["原則5", "選択肢を出せ", "「延命 vs 売却 vs 撤去」など判断材料を並べる"],
    ["原則6", "信頼を証明しろ", "作業前後の写真・データ比較で効果を見える化する"],
]
headers_pr = ["No", "原則名", "内容"]
for ci, h in enumerate(headers_pr, 1):
    set_cell(ws1, r, ci, h, header_font, header_fill, center_align, thin_border)
r += 1
for p in principles:
    for ci, v in enumerate(p, 1):
        set_cell(ws1, r, ci, v, body_font, alignment=wrap_align, border=thin_border)
    r += 1

# ====================================================================
# Sheet 2: Best 30 Ideas
# ====================================================================
ws2 = wb.create_sheet("2_昇華アイデアBEST30")
ws2.sheet_properties.tabColor = "2E75B6"

col_widths_2 = [6, 10, 10, 30, 35, 45, 45, 12]
for ci, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

r = 1
set_cell(ws2, r, 1, "昇華アイデア BEST 30 - 6テーマ構成", title_font, title_fill, center_align, merge_end_col=8)
r += 2

# Theme descriptions
themes_desc = [
    ("A", "金額換算で行動を変える（5件）", "見える化を「金額インパクト」に変換し、顧客の意思決定を加速"),
    ("B", "比較で気づかせる（5件）", "他の発電所や業界平均との比較で、自分の立ち位置を認識させる"),
    ("C", "次のアクションを自動提示（5件）", "アラートの先にある「具体的にどうすればいいか」まで踏み込む"),
    ("D", "お金の流れを経営に直結（5件）", "発電データとP/L・CF・税務をつなげ、経営判断の材料にする"),
    ("E", "出口・延命の選択肢を出す（5件）", "設備の将来について、データに基づく選択肢を並べて比較できる"),
    ("F", "信頼と透明性の証明（5件）", "メンテ業者への不信感を解消し、作業の価値を数字で証明する"),
]

for theme_code, theme_name, theme_desc in themes_desc:
    set_cell(ws2, r, 1, theme_code, Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF"),
             PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"), center_align, thin_border)
    set_cell(ws2, r, 2, theme_name, h3_font, alignment=wrap_align, merge_end_col=4)
    set_cell(ws2, r, 5, theme_desc, body_font_s, alignment=wrap_align, merge_end_col=8)
    r += 1

r += 1

# Table Headers
headers_2 = ["No", "テーマ", "元No.", "アイデア名", "サブタイトル", "顧客の本当のペイン", "昇華後のシステム設計", "Phase"]
for ci, h in enumerate(headers_2, 1):
    set_cell(ws2, r, ci, h, header_font, header_fill, center_align, thin_border)
r += 1

# 30 Ideas Data
ideas_30 = [
    # Theme A
    ["A1", "A", "1,3", "損失金額リアルタイムアラート", "kWhの異常を「月○万円の損失」で通知", "P1: 数字が専門的で異常かどうかわからない", "PR値低下をJEPX単価で金額換算し、LINE/メールでプッシュ通知。過去3ヶ月比較グラフ付き", "Phase1"],
    ["A2", "A", "5,7", "パネル劣化 損益シミュレーター", "交換費用 vs 放置損失を自動比較", "P1: 劣化しているのはわかるが、いつ交換すべきかわからない", "年次劣化率から将来損失を予測。パネル交換ROIを3パターン（即時/1年後/3年後）で提示", "Phase1"],
    ["A3", "A", "12", "天候連動 発電予実ギャップ通知", "「晴れなのに発電が低い」を金額で警告", "P1: 天気が悪いのか設備が悪いのか区別できない", "気象APIと発電データを突合。天候補正後の「本来の発電量」との差分を金額表示", "Phase1"],
    ["A4", "A", "15,18", "JEPX連動 売電タイミング最適化", "30分値で「いつ売ると最も得か」を提示", "P5: 売電収入が市場価格でどう変わるか把握できない", "JEPXスポット価格×発電予測で、蓄電池充放電の最適スケジュールを自動生成", "Phase2"],
    ["A5", "A", "22", "年間キャッシュフロー予測レポート", "FIT残存期間の累積収支を年次で予測", "P5: 投資回収がいつ終わるのかわからない", "FIT単価×発電予測−維持費−税金で20年キャッシュフロー表を自動生成。IRR/NPV算出", "Phase2"],

    # Theme B
    ["B1", "B", "2,6", "発電所ベンチマークレポート", "同地域・同規模の発電所と自動比較", "P1: 自分の発電所が平均的なのか悪いのかわからない", "地域・容量・パネルメーカーで類似発電所をグルーピング。偏差値方式でランク表示", "Phase1"],
    ["B2", "B", "8,10", "メンテナンス費用 業界水準比較", "あなたの保守費は業界平均の○○%", "P2: メンテ費用の妥当性が判断できない", "匿名化された保守費データベースと比較。kWあたりの保守単価ランキング表示", "Phase2"],
    ["B3", "B", "25", "除草コスト最適化マップ", "除草頻度×費用を近隣発電所と比較", "P2: 除草業者の請求が高いのか安いのかわからない", "衛星画像の草丈推定+地域除草単価DBで、最適除草頻度と年間コストを算出", "Phase2"],
    ["B4", "B", "30", "PCS効率 メーカー横断比較", "あなたのPCS効率は同メーカー内で何位？", "P1: PCSの性能低下を客観的に把握できない", "メーカー・型番別の効率分布を統計分析。経年劣化カーブとの乖離でアラート", "Phase2"],
    ["B5", "B", "35,40", "故障率ヒートマップ 全国版", "日本地図上で故障リスクを色分け表示", "P1: 自分の地域特有のリスクがわからない", "塩害・積雪・風速・日照データと故障実績を重畳。地域別リスクスコア算出", "Phase3"],

    # Theme C
    ["C1", "C", "4,9", "ワンタップ駆けつけ手配", "異常検知→業者手配→見積もりまで自動化", "P2: 異常を知っても、どの業者に頼めばいいかわからない", "アラート画面から1タップで地域の認定業者に見積依頼。3社比較で即発注可能", "Phase1"],
    ["C2", "C", "11,14", "法令改正 影響度自動診断", "電気事業法/FIT法改正の影響を自動チェック", "P3: 法改正のニュースは見るが、自分に関係あるかわからない", "発電所スペック×法令DBのマッチングで、影響ありの項目のみ抽出。対応ToDoリスト生成", "Phase2"],
    ["C3", "C", "20,23", "保険最適化シミュレーター", "設備リスクに応じた保険プランを自動提案", "P3: 保険の過不足がわからず、何となく継続している", "発電所の災害リスクスコア×設備経年×過去事故率で、最適な保険カバレッジを算出", "Phase3"],
    ["C4", "C", "28,32", "パネル洗浄 効果予測AI", "「洗浄すると月○万円回復」を事前に算出", "P2: 洗浄の効果がわからず、投資すべきか迷う", "汚損度推定（衛星+ドローン）×過去洗浄効果データで回復発電量を予測。ROI付き提案", "Phase2"],
    ["C5", "C", "45,50", "蓄電池導入 即時シミュレーション", "「蓄電池を入れたら年間○万円得」を30秒で算出", "P4: 蓄電池の投資対効果がすぐにわからない", "発電パターン×JEPX価格×電力需要パターンで最適容量と回収年数を即算出", "Phase2"],

    # Theme D
    ["D1", "D", "33,37", "経営ダッシュボード連携", "発電データを月次P/Lに自動変換", "P5: 発電量は見ているが、利益にどう影響するか不明", "売電収入−変動費−固定費を自動計算。会計ソフトAPI連携でP/L項目に自動マッピング", "Phase2"],
    ["D2", "D", "38,42", "減価償却 最適化アドバイザー", "「今年の税負担を○万円軽減できます」", "P5: 設備の税務処理が最適かどうかわからない", "即時償却/定率法/定額法のシミュレーション。税理士への相談用レポート自動生成", "Phase3"],
    ["D3", "D", "46,48", "補助金マッチング自動通知", "あなたの設備に使える補助金を自動検索", "P3: 補助金があるらしいが、調べる時間がない", "設備情報×補助金DBの定期マッチング。申請期限リマインダー+申請書テンプレート生成", "Phase2"],
    ["D4", "D", "52,55", "融資条件 最適化レポート", "発電実績データで融資借り換えを支援", "P5: 発電実績データを融資交渉に活かせていない", "過去発電実績のエビデンスレポート自動生成。金融機関向けフォーマットで出力", "Phase3"],
    ["D5", "D", "58,60", "法人税 節税シミュレーション", "太陽光設備の税制優遇を最大活用", "P5: 税制優遇の活用余地がわからない", "中小企業投資促進税制等の適用可否を自動判定。税効果金額を試算して税理士レポート化", "Phase3"],

    # Theme E
    ["E1", "E", "61,65", "設備価値リアルタイム査定", "「今売ったらいくら？」を常に更新", "P4: 設備を売りたいが、適正価格がわからない", "発電実績×設備年齢×FIT残年数×中古市場データで時価を算出。月次更新", "Phase2"],
    ["E2", "E", "67,70", "延命 vs 撤去 vs 売却 比較表", "3つの選択肢をNPVで比較", "P4: FIT終了後の設備をどうすべきか判断できない", "各シナリオの20年NPVを自動計算。感度分析（電気料金上昇率、劣化率）付き", "Phase2"],
    ["E3", "E", "72,75", "リパワリング経済性診断", "「パネル交換で年○万円増収」を自動試算", "P4: リパワリングの投資対効果が不明", "現行パネル効率vs最新パネル効率×残FIT期間で増収額を算出。メーカー見積連携", "Phase3"],
    ["E4", "E", "78,80", "卒FIT 自家消費切替シミュレーション", "「自家消費に切り替えると年○万円お得」", "P4: FIT終了後に全量売電を続けるべきか迷う", "電力需要パターン×発電パターン×電気料金単価で自家消費メリットを試算", "Phase2"],
    ["E5", "E", "82", "カーポート太陽光+EV 統合提案", "駐車場→発電所+EV充電スポット化", "P4: 遊休資産（駐車場）の活用方法がわからない", "カーポート太陽光+EV充電+V2Bの統合経済性シミュレーション。投資回収年数付き", "Phase3"],

    # Theme F
    ["F1", "F", "85,88", "メンテナンス作業 Before/After証明書", "作業効果を写真+データで証明", "P2: メンテ業者が本当に作業したのか確認できない", "作業前後のIVカーブ測定+写真+発電量比較を自動レポート化。改善金額を明示", "Phase1"],
    ["F2", "F", "90,92", "業者評価 見える化スコアカード", "メンテ業者を5段階で評価・比較", "P2: 業者の品質を客観的に評価する基準がない", "対応速度・復旧率・顧客満足度・再発率で業者をスコアリング。エリア別ランキング", "Phase2"],
    ["F3", "F", "94,95", "24時間監視 安心ステータス表示", "「あなたの発電所は今、正常です」を常に表示", "P1: 異常がないときも不安で、何度もログインしてしまう", "正常時も「ステータス:正常/最終チェック:3分前」を常時表示。異常時のみ詳細展開", "Phase1"],
    ["F4", "F", "97", "発電所 健康診断レポート（年次）", "人間ドックのように年1回の総合レポート", "P1: 日々のデータは見ているが、全体的な健康状態がわからない", "年間発電量・PR値推移・故障回数・メンテ費用の総合スコア。前年比・業界比較付き", "Phase2"],
    ["F5", "F", "100", "オーナーコミュニティ 匿名相談", "同じ悩みを持つオーナー同士で情報交換", "P2: 業者以外に相談できる相手がいない", "匿名掲示板+AI要約で、地域・設備別の課題と解決策をナレッジ化。専門家回答も統合", "Phase3"],
]

for idea in ideas_30:
    theme_code = idea[1]
    fill = theme_fills.get(theme_code, None)
    for ci, v in enumerate(idea, 1):
        f = fill if ci <= 2 else None
        set_cell(ws2, r, ci, v, body_font, f, wrap_align, thin_border)
    r += 1

# ====================================================================
# Sheet 3: New Ideas + Combinations
# ====================================================================
ws3 = wb.create_sheet("3_新規アイデア＋掛け合わせ")
ws3.sheet_properties.tabColor = "ED7D31"

col_widths_3 = [6, 30, 40, 45, 45, 12]
for ci, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

r = 1
set_cell(ws3, r, 1, "新規アイデア + 掛け合わせアイデア", title_font, title_fill, center_align, merge_end_col=6)
r += 2

# New Ideas
set_cell(ws3, r, 1, "■ 元の100本には無かった新規アイデア（5件）", h2_font)
r += 1
headers_3n = ["No", "アイデア名", "概要", "顧客ペイン", "システム設計", "Phase"]
for ci, h in enumerate(headers_3n, 1):
    set_cell(ws3, r, ci, h, header_font, header_fill, center_align, thin_border)
r += 1

new_ideas = [
    ["N1", "発電所オーナー専用 経営相談AI", "発電・税務・法務の横断相談にAIが24時間対応", "P2,P3: 専門知識がなく、誰に聞けばいいかわからない", "RAG(検索拡張生成)で法令DB+税務DB+発電データを統合。質問をカテゴリ分類して専門家にエスカレーション", "Phase2"],
    ["N2", "災害リスク事前通知＋自動保全モード", "台風・豪雨の48時間前に設備保全アクションを指示", "P3: 災害対策が後手に回り、被害が拡大する", "気象警報API×設備脆弱性DBで事前リスク評価。PCSシャットダウン・排水ポンプ作動などの自動保全シナリオ実行", "Phase3"],
    ["N3", "発電所 承継サポートパッケージ", "オーナーの高齢化・相続に伴う事業承継を総合支援", "P4: 自分に何かあった時、発電所をどうすればいいかわからない", "設備評価+法的手続きガイド+買い手マッチング+管理移行サポートをワンストップ提供", "Phase3"],
    ["N4", "電力需要家マッチング プラットフォーム", "発電オーナーと電力消費企業を直接つなぐ", "P4: 卒FIT後の売電先が見つからない", "需要家の消費パターンと発電パターンをマッチング。PPA契約の条件自動生成", "Phase3"],
    ["N5", "設備メーカー保証 自動請求アシスト", "メーカー保証の請求漏れを自動検出・申請支援", "P2: 保証期間内の不具合でも請求方法がわからず泣き寝入り", "設備型番×保証条件DB×故障データで保証請求可能な案件を自動抽出。申請書テンプレート+エビデンス自動生成", "Phase2"],
]
for idea in new_ideas:
    for ci, v in enumerate(idea, 1):
        f = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if ci == 1 else None
        set_cell(ws3, r, ci, v, body_font, f, wrap_align, thin_border)
    r += 1

r += 2

# Combination Ideas
set_cell(ws3, r, 1, "■ 既存アイデアの掛け合わせ（5件）", h2_font)
r += 1
headers_3c = ["No", "アイデア名", "掛け合わせ元", "シナジー効果", "システム設計", "Phase"]
for ci, h in enumerate(headers_3c, 1):
    set_cell(ws3, r, ci, h, header_font, header_fill, center_align, thin_border)
r += 1

combo_ideas = [
    ["X1", "金額換算×ベンチマーク 統合ダッシュボード", "A1 + B1", "損失金額を業界比較で表示し、「あなたは業界平均より月○万円損している」と伝える", "PR値金額換算+地域ベンチマークを1画面に統合。改善余地を金額ランキングで表示", "Phase1"],
    ["X2", "法改正影響×経営P/L 連動アラート", "C2 + D1", "法改正が自社P/Lに与える影響額を即座に算出", "法令改正DB×設備情報×P/Lデータの3層マッチング。影響額の大きい改正のみ通知", "Phase2"],
    ["X3", "設備査定×出口比較×承継サポート 統合", "E1 + E2 + N3", "設備の現在価値から最適出口戦略を提案し、承継まで一気通貫", "リアルタイム査定→3シナリオNPV比較→選択に応じた承継/売却/撤去手続きガイド", "Phase3"],
    ["X4", "メンテ証明×業者評価×ワンタップ手配", "F1 + F2 + C1", "作業品質の証明→業者評価→次回発注を一気通貫で実現", "作業証明書の品質スコア→業者評価DB更新→次回異常時の業者自動レコメンド", "Phase2"],
    ["X5", "EV統合×電力マッチング×JEPX最適化", "E5 + N4 + A4", "カーポート太陽光+EV+需要家マッチングで電力の地産地消を実現", "発電予測×EV充電需要×近隣需要家の消費パターンで最適な電力フローを自動制御", "Phase3"],
]
for idea in combo_ideas:
    for ci, v in enumerate(idea, 1):
        f = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") if ci == 1 else None
        set_cell(ws3, r, ci, v, body_font, f, wrap_align, thin_border)
    r += 1

# ====================================================================
# Sheet 4: Implementation Roadmap
# ====================================================================
ws4 = wb.create_sheet("4_実装ロードマップ")
ws4.sheet_properties.tabColor = "70AD47"

col_widths_4 = [10, 20, 15, 45, 35, 25]
for ci, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(ci)].width = w

r = 1
set_cell(ws4, r, 1, "実装ロードマップ + 共通基盤", title_font, title_fill, center_align, merge_end_col=6)
r += 2

set_cell(ws4, r, 1, "■ 5フェーズ実装ロードマップ", h2_font)
r += 1
headers_4 = ["Phase", "期間", "対象アイデア数", "主要施策", "必要技術", "KPI"]
for ci, h in enumerate(headers_4, 1):
    set_cell(ws4, r, ci, h, header_font, header_fill, center_align, thin_border)
r += 1

phases = [
    ["Phase 1", "0-6ヶ月", "8件", "金額換算アラート、Before/After証明、ベンチマーク基礎、安心ステータス表示", "JEPX API連携、PR値計算エンジン、画像管理、プッシュ通知", "MAU 50%向上、アラート対応率80%"],
    ["Phase 2", "6-12ヶ月", "14件", "経営ダッシュボード、業者評価、法改正診断、出口戦略比較、蓄電池シミュ", "会計API連携、業者DB、法令DB、NPV計算エンジン、気象API", "有料プラン転換率30%、顧客LTV 2倍"],
    ["Phase 3", "12-18ヶ月", "10件", "災害事前通知、リパワリング診断、減価償却最適化、カーポート+EV統合", "気象警報API、衛星画像解析、税務シミュレーション、EV充電制御", "新規顧客獲得率50%向上"],
    ["Phase 4", "18-24ヶ月", "5件", "電力マッチング、承継サポート、AI経営相談、全国故障ヒートマップ", "マッチングエンジン、RAG/LLM、GIS統合、法務DB", "プラットフォーム収入 月額○万円"],
    ["Phase 5", "24-36ヶ月", "3件", "掛け合わせ統合アイデア（X3, X5等）の本格展開", "マイクロサービス統合、API Gateway、統合ダッシュボード", "エコシステム参加企業100社"],
]

phase_fills = [
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    PatternFill(start_color="E2D9F3", end_color="E2D9F3", fill_type="solid"),
    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
]

for pi, phase in enumerate(phases):
    for ci, v in enumerate(phase, 1):
        f = phase_fills[pi] if ci == 1 else None
        set_cell(ws4, r, ci, v, body_font, f, wrap_align, thin_border)
    r += 1

r += 2

# Infrastructure Platforms
set_cell(ws4, r, 1, "■ 4つの共通基盤プラットフォーム", h2_font)
r += 1
headers_4p = ["No", "プラットフォーム名", "役割", "対応アイデア", "技術スタック"]
for ci, h in enumerate(headers_4p, 1):
    set_cell(ws4, r, ci, h, header_font, header_fill, center_align, thin_border)
r += 1

platforms = [
    ["PF1", "金額換算エンジン", "kWh/PR値を円に変換する共通API。JEPX連動で常に最新単価", "A1-A5, B1, X1", "JEPX API, Redis Cache, RESTful API"],
    ["PF2", "イベントバス (通知基盤)", "異常検知→判定→通知→アクション提示の一気通貫パイプライン", "A1, C1-C5, F3", "Apache Kafka, WebSocket, LINE/Slack Bot"],
    ["PF3", "Customer 360 (顧客統合DB)", "発電データ+経営データ+設備情報+取引履歴を統合した顧客プロファイル", "D1-D5, E1-E5, N1", "PostgreSQL, GraphQL, Data Lake"],
    ["PF4", "デジタルツイン (設備シミュレーション)", "設備の仮想モデルで、交換/延命/撤去のシナリオをシミュレーション", "E1-E5, E5(EV), N2", "Unity/Three.js, ML予測モデル, IoTデータ"],
]

infra_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
for pf in platforms:
    for ci, v in enumerate(pf, 1):
        f = infra_fill if ci == 1 else None
        set_cell(ws4, r, ci, v, body_font, f, wrap_align, thin_border)
    r += 1

# Save
wb.save(output_path)
print(f"Excel saved to: {output_path}")
