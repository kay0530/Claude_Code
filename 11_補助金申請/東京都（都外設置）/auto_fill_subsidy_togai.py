"""
Tokyo Subsidy Application Form Auto-Fill Script - Outside-Tokyo Installation (都外設置)
Automatically fills in solar power generation data into the official Excel form.

Target case: 名糖運輸（日高物流センター）- 埼玉県日高市
Environmental value location: 西東京物流センター - 東京都青梅市

Target sheets (4 sheets):
  1. 共通１－２(太陽光）          - Solar equipment specs + monthly generation/consumption
  2. 共通１－４（環境価値）        - Environmental value utilization (NEW for togai)
  3. 共通２_全体                  - Application type + cost summary
  4. 共通２_太陽光発電            - Cost detail transcription
"""

import openpyxl
import os
import sys
import warnings

warnings.filterwarnings("ignore", category=UserWarning)

# File paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, "3_r7chisan_togai_yoshiki_20260106 .xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "r7chisan_togai_yoshiki_auto_filled.xlsx")


def is_formula(cell):
    """Check if a cell contains a formula."""
    return isinstance(cell.value, str) and cell.value.startswith("=")


def safe_write(ws, cell_ref, value, description=""):
    """Write value to cell only if it does not contain a formula.
    Handles merged cells by writing to the top-left cell of the merge range.
    """
    from openpyxl.cell.cell import MergedCell
    cell = ws[cell_ref]
    # If the cell is a MergedCell, find the top-left cell of the merge range
    if isinstance(cell, MergedCell):
        for merge_range in ws.merged_cells.ranges:
            if cell.coordinate in merge_range:
                top_left = str(merge_range).split(":")[0]
                cell = ws[top_left]
                print(f"  [MERGE] {cell_ref} -> writing to {top_left} (merged range: {merge_range})")
                break
    if is_formula(cell):
        print(f"  [SKIP] {cell_ref}: formula detected ({str(cell.value)[:50]}...) - {description}")
        return False
    else:
        cell.value = value
        print(f"  [WRITE] {cell_ref} = {repr(value)} - {description}")
        return True


def find_sheet_by_name(wb, keywords):
    """Find a sheet by matching keywords against sheet names.

    Returns (index, worksheet) or raises ValueError if not found.
    Searches for sheets whose name contains ANY of the given keywords.
    """
    for i, name in enumerate(wb.sheetnames):
        for kw in keywords:
            if kw in name:
                return i, wb.worksheets[i]
    raise ValueError(f"Sheet not found matching keywords: {keywords}\n"
                     f"Available sheets: {wb.sheetnames}")


# =============================================================================
# DATA - 名糖運輸 日高物流センター (Installation site: 埼玉県日高市)
# =============================================================================

# Solar module data
MODULE_MANUFACTURER = "ジンコソーラージャパン"
MODULE_MODEL = "JKM585N-72HL4-V-J"
MODULE_OUTPUT_W = 585
MODULE_COUNT = 651
# DC capacity = 585 * 651 / 1000 = 380.835 → 380.83 kW (formula)

# Power conditioner data
PCS_MANUFACTURER = "ファーウェイジャパン"
PCS_MODEL = "SUN2000-50KTL-JPM0"
PCS_RATED_OUTPUT_KW = 50
PCS_UNIT_COUNT = 5
# AC capacity = 50 * 5 = 250.00 kW (formula)

# Monthly consumption at installation site (日高物流センター, kWh)
CONSUMPTION_HIDAKA = {
    "apr": 209942, "may": 211399, "jun": 244121,
    "jul": 301642, "aug": 302525, "sep": 282069,
    "oct": 240219, "nov": 205582, "dec": 198417,
    "jan": 189029, "feb": 171924, "mar": 197362,
}
ANNUAL_CONSUMPTION_HIDAKA = sum(CONSUMPTION_HIDAKA.values())  # 2,754,231

# Monthly generation from SolarPro simulation (kWh)
GENERATION_SOLARPRO = {
    "apr": 45593.48, "may": 50237.18, "jun": 42226.08,
    "jul": 46111.67, "aug": 43626.77, "sep": 35713.80,
    "oct": 29410.80, "nov": 24250.90, "dec": 24126.82,
    "jan": 28992.27, "feb": 30952.83, "mar": 39130.71,
}
ANNUAL_GENERATION = sum(GENERATION_SOLARPRO.values())  # 440,373.31

# Monthly consumption at environmental value location (西東京物流センター, kWh)
CONSUMPTION_NISHITOKYO = {
    "apr": 63499, "may": 69616, "jun": 78477,
    "jul": 93231, "aug": 109141, "sep": 108003,
    "oct": 92895, "nov": 75448, "dec": 63814,
    "jan": 63866, "feb": 61594, "mar": 55819,
}
ANNUAL_CONSUMPTION_NISHITOKYO = sum(CONSUMPTION_NISHITOKYO.values())  # 935,403

# Environmental value data
ENV_COMPANY_NAME = "名糖運輸株式会社"
ENV_FACILITY_NAME = "西東京物流センター"
ENV_CONTRACTOR_NAME = "名糖運輸株式会社"
ENV_ADDRESS = "東京都　青梅市　新町　８丁目１－１９"
ENV_ENERGY_TYPE = "電力"
ENV_SUBSIDY_RATE = "1/2"
# Certificate amount = ANNUAL_GENERATION * ENV_SUBSIDY_RATE = 440373 * 1/2 ≈ 220,187
ENV_CERTIFICATE_AMOUNT = 220187
ENV_CERTIFICATE_TYPE = "グリーン電力証書"
ENV_CERTIFICATE_ISSUER = "証書発行事業者：デジタルグリッド株式会社"

# Environmental value schedule text
ENV_SCHEDULE_LINES = [
    (117, "令和〇年〇月　設備の審査を審査機関（一般財団法人　日本品質保証機構）へ"),
    (124, "                  審査依頼予定"),
    (125, "令和〇年〇月　設備審査終了予定"),
    (126, "令和〇年〇月　発電量計測開始"),
    (127, "                   以降令和〇年〇月まで継続して発電量を計測予定"),
    (128, "令和〇年〇月　グリーン電力証書発行を行い都内事業所で証書利用開始予定"),
    (129, "　　　　　　　　　以降毎年〇月にグリーン電力証書発行を行い、"),
    (130, "　　　　　　　　　令和〇年〇月まで証書を都内事業所で利用予定"),
    (132, "証書発行事業者：デジタルグリッド株式会社"),
]

# Equipment cost details for 日高物流センター
# Format: (row, detail_no, item_name, subsidy_class, cost_category,
#          amount, is_non_target)
# Rows 13-36 (continuous, no gap rows in template)
# Template columns: B=費目区分, C=品目名, D=明細番号, E=助成区分, F=見積金額
# I column = formula (skip all writes)
EQUIPMENT_ITEMS = [
    # Equipment cost items (設備費) - rows 13-24
    (13, "A-1", "太陽電池モジュール", "発電設備", "設備費", 8723400, False),
    (14, "A-2", "パワーコンディショナ", "発電設備", "設備費", 4065000, False),
    (15, "A-3", "太陽電池モジュール用架台", "発電設備", "設備費", 1390080, False),
    (16, "A-4", "ダウントランス", "発電設備", "設備費", 3200000, False),
    (17, "A-5", "データロガー", "発電設備", "設備費", 82000, True),
    (18, "A-6", "監視計測装置", "発電設備", "設備費", 489000, False),
    (19, "A-6", "監視計測装置運用費", "発電設備", "設備費", 720000, True),
    (20, "A-7", "OVGR+RPR", "発電設備", "設備費", 107320, False),
    (21, "A-8", "ZPD", "発電設備", "設備費", 75000, False),
    (22, "A-9", "検定付計量器", "発電設備", "設備費", 500000, False),
    (23, "A-10", "太陽電池モジュール延長ケーブル", "発電設備", "設備費", 1334760, False),
    (24, "A-11", "PCS/トランス取付架台", "発電設備", "設備費", 350000, False),
    (25, "A-12", "電気材料・配線資材", "発電設備", "設備費", 2833680, False),
    # Construction cost items (工事費) - rows 26-36
    (26, "B-1,2,3", "ソーラー金具、太陽電池モジュール、盤類据付工事", "発電設備", "工事費", 3780000, False),
    (27, "B-4", "電工費", "発電設備", "工事費", 6550000, False),
    (28, "B-5", "高圧設備改造工事", "発電設備", "工事費", 810000, False),
    (29, "B-6", "揚重費", "発電設備", "工事費", 640000, False),
    (30, "B-7", "資材運送費・搬入費", "発電設備", "工事費", 270000, False),
    (31, "B-8", "仮設費", "発電設備", "工事費", 742000, False),
    (32, "B-9", "安全対策費", "発電設備", "工事費", 640000, False),
    (33, "B-10", "試運転調整", "発電設備", "工事費", 512000, False),
    (34, "B-11", "使用前自己確認検査費用", "発電設備", "工事費", 516000, False),
    (35, "B-12", "産廃費", "発電設備", "工事費", 136000, False),
    (36, "B-13", "諸経費", "発電設備", "工事費", 1350000, False),
]

TOTAL_COST = sum(a for _, _, _, _, _, a, _ in EQUIPMENT_ITEMS)  # 39,816,240
NON_TARGET_COST = sum(a for _, _, _, _, _, a, nt in EQUIPMENT_ITEMS if nt)  # 802,000
TARGET_COST = TOTAL_COST - NON_TARGET_COST  # 39,014,240


# =============================================================================
# FILL FUNCTIONS
# =============================================================================

def fill_solar_sheet(wb):
    """Fill 共通１－２(太陽光) - Solar equipment specs + monthly generation/consumption.

    Cell mapping (verified from template openpyxl analysis):
      G6  = system output (formula: MIN of DC, AC) - SKIP
      G12 = module manufacturer (G12:I12 merged)
      G13 = module model (G13:I13 merged)
      G15 = module output W (G15:I15 merged)
      G16 = module count (G16:I16 merged)
      G17 = DC capacity (formula) - SKIP
      G39 = PCS manufacturer (G39:I39 merged)
      G40 = PCS model (G40:I40 merged)
      G42 = PCS rated output per unit kW (G42:I42 merged)
      G43 = PCS unit count (G43:I43 merged)
      G44 = total AC capacity (formula) - SKIP
      E126-J126 = monthly consumption Apr-Sep (cols E-J, not F-K)
      E127-J127 = monthly generation Apr-Sep
      E130-J130 = monthly consumption Oct-Mar
      E131-J131 = monthly generation Oct-Mar
      G134 = annual consumption (formula) - SKIP
      G136 = annual generation (formula) - SKIP
      G138 = annual difference (formula) - SKIP
    """
    _, ws = find_sheet_by_name(wb, ["共通１－２(太陽光", "共通1-2(太陽光"])
    print(f"\n{'='*60}")
    print(f"Sheet: {ws.title}")
    print(f"{'='*60}")

    # ------- Solar module (rows 11-17) -------
    print(f"\n--- (1) Solar module ---")
    print(f"  [INFO] G6 contains formula (system output) - auto-calculated")

    safe_write(ws, "G11", "Ａ-１", "Module quotation detail number (見積明細番号)")
    safe_write(ws, "G12", MODULE_MANUFACTURER, "Module manufacturer")
    safe_write(ws, "G13", MODULE_MODEL, "Module model")
    safe_write(ws, "G15", MODULE_OUTPUT_W, "Module output (W)")
    safe_write(ws, "G16", MODULE_COUNT, "Module count")
    print(f"  [INFO] G17 contains formula (DC capacity) - expected: 380.83 kW")

    # ------- Power conditioner (rows 38-44) -------
    print(f"\n--- (2) Power conditioner ---")
    safe_write(ws, "G38", "Ａ-２", "PCS quotation detail number (見積明細番号)")
    safe_write(ws, "G39", PCS_MANUFACTURER, "PCS manufacturer")
    safe_write(ws, "G40", PCS_MODEL, "PCS model")
    safe_write(ws, "G42", PCS_RATED_OUTPUT_KW, "PCS rated output per unit (kW)")
    safe_write(ws, "G43", PCS_UNIT_COUNT, "PCS unit count")
    print(f"  [INFO] G44 contains formula (total AC) - expected: 250.00 kW")

    # ------- Monthly consumption (installation site: 日高) -------
    print(f"\n--- (3) Monthly power consumption (日高物流センター) ---")

    # Apr-Sep: row 126, columns E-J
    months_apr_sep = ["apr", "may", "jun", "jul", "aug", "sep"]
    cols_ej = ["E", "F", "G", "H", "I", "J"]
    for col, month in zip(cols_ej, months_apr_sep):
        safe_write(ws, f"{col}126", CONSUMPTION_HIDAKA[month],
                   f"{month.capitalize()} consumption")

    # Oct-Mar: row 130, columns E-J
    months_oct_mar = ["oct", "nov", "dec", "jan", "feb", "mar"]
    for col, month in zip(cols_ej, months_oct_mar):
        safe_write(ws, f"{col}130", CONSUMPTION_HIDAKA[month],
                   f"{month.capitalize()} consumption")

    # ------- Monthly generation (SolarPro) -------
    print(f"\n--- (4) Monthly power generation (SolarPro simulation) ---")

    # Apr-Sep: row 127, columns E-J
    for col, month in zip(cols_ej, months_apr_sep):
        safe_write(ws, f"{col}127", GENERATION_SOLARPRO[month],
                   f"{month.capitalize()} generation (SolarPro)")

    # Oct-Mar: row 131, columns E-J
    for col, month in zip(cols_ej, months_oct_mar):
        safe_write(ws, f"{col}131", GENERATION_SOLARPRO[month],
                   f"{month.capitalize()} generation (SolarPro)")

    # ------- Annual totals (formulas) -------
    print(f"\n--- (5) Annual totals ---")
    print(f"  [INFO] G134 formula (annual consumption) - expected: {ANNUAL_CONSUMPTION_HIDAKA:,}")
    print(f"  [INFO] G136 formula (annual generation) - expected: {ANNUAL_GENERATION:,.2f}")
    print(f"  [INFO] G138 formula (difference) - expected: {ANNUAL_CONSUMPTION_HIDAKA - ANNUAL_GENERATION:,.2f}")

    # Verification
    gen_total = sum(GENERATION_SOLARPRO.values())
    con_total = sum(CONSUMPTION_HIDAKA.values())
    print(f"  [VERIFY] Consumption total: {con_total:,} kWh")
    print(f"  [VERIFY] Generation total: {gen_total:,.2f} kWh")


def fill_env_value_sheet(wb):
    """Fill 共通１－４（環境価値） - Environmental value utilization plan.

    This sheet is NEW for 都外設置 (not present in 都内設置).
    It documents where the environmental value (グリーン電力証書) is utilized
    within Tokyo, even though the solar installation is outside Tokyo.

    Cell mapping (verified from template openpyxl analysis):
      Location 1 header:
        E3  = company name (E3:G3 merged)
        I3  = energy type (I3:J3 merged, dropdown: 電力/ガス)
        E4  = address (E4:J4 merged) - facility address
        B5:D8 = label "年間電力消費量" (merged, DO NOT WRITE)
      Location 1 monthly consumption:
        E6-J6  = consumption Apr-Sep (cols E-J)
        E8-J8  = consumption Oct-Mar (cols E-J)
        J9     = annual total (formula) - SKIP
      Summary (lower area):
        H95  = annual consumption total (formula) - SKIP
        H97  = subsidy rate (H97:I97 merged)
        H99  = certificate amount (formula) - SKIP
        H102 = certificate type (H102:I102 merged)
        D105 = schedule text (D105:J117 giant merged cell - write as single text)
    """
    _, ws = find_sheet_by_name(wb, ["共通１－４", "共通1-4", "環境価値"])
    print(f"\n{'='*60}")
    print(f"Sheet: {ws.title}")
    print(f"{'='*60}")

    # ------- Header / Location 1 basic info -------
    print(f"\n--- (1) Environmental value location info ---")
    safe_write(ws, "E3", ENV_COMPANY_NAME, "Company name (E3:G3 merged)")
    safe_write(ws, "I3", ENV_CONTRACTOR_NAME, "Power contractor name (I3:J3 merged)")
    safe_write(ws, "E4", ENV_ADDRESS, "Address (E4:J4 merged)")
    # Note: B5:D8 is a merged label cell "年間電力消費量" - DO NOT WRITE
    # Note: C4:D4 is a label "住所" - DO NOT WRITE

    # ------- Monthly consumption at 西東京 (利用場所1) -------
    print(f"\n--- (2) Monthly consumption (西東京物流センター) ---")

    months_apr_sep = ["apr", "may", "jun", "jul", "aug", "sep"]
    months_oct_mar = ["oct", "nov", "dec", "jan", "feb", "mar"]
    cols_ej = ["E", "F", "G", "H", "I", "J"]

    # Apr-Sep: row 6, columns E-J
    for col, month in zip(cols_ej, months_apr_sep):
        safe_write(ws, f"{col}6", CONSUMPTION_NISHITOKYO[month],
                   f"{month.capitalize()} consumption (西東京)")

    # Oct-Mar: row 8, columns E-J
    for col, month in zip(cols_ej, months_oct_mar):
        safe_write(ws, f"{col}8", CONSUMPTION_NISHITOKYO[month],
                   f"{month.capitalize()} consumption (西東京)")

    print(f"  [INFO] J9 formula (annual subtotal) - expected: {ANNUAL_CONSUMPTION_NISHITOKYO:,}")

    # ------- Summary area -------
    print(f"\n--- (3) Certificate summary ---")
    print(f"  [INFO] H95 formula (annual consumption total) - SKIP")
    safe_write(ws, "H97", ENV_SUBSIDY_RATE, "Subsidy rate (H97:I97 merged)")
    print(f"  [INFO] H99 formula (certificate amount) - SKIP")
    safe_write(ws, "H102", ENV_CERTIFICATE_TYPE, "Certificate type (H102:I102 merged)")

    # ------- Schedule text (D105:J117 = ONE giant merged cell) -------
    print(f"\n--- (4) Environmental value utilization schedule ---")
    # Combine all schedule lines into a single text with newlines
    schedule_text = "\n".join(text for _, text in ENV_SCHEDULE_LINES)
    safe_write(ws, "D105", schedule_text, "Schedule text (D105:J117 merged cell)")

    print(f"\n  [VERIFY] Annual consumption (西東京): {ANNUAL_CONSUMPTION_NISHITOKYO:,} kWh")
    print(f"  [VERIFY] Certificate amount: {ENV_CERTIFICATE_AMOUNT:,} kWh")


def fill_overview_sheet(wb):
    """Fill 共通２_全体 - Application type + cost summary.

    Cell mapping (verified from template openpyxl analysis):
      D2  = application type dropdown (D2:F2 merged)

    Most cells are formulas that reference 共通２_太陽光発電.
    Only D2 needs manual input.
    """
    _, ws = find_sheet_by_name(wb, ["共通２_全体", "共通2_全体"])
    print(f"\n{'='*60}")
    print(f"Sheet: {ws.title}")
    print(f"{'='*60}")

    # ------- Application type dropdown -------
    print(f"\n--- (1) Application type dropdown ---")
    safe_write(ws, "D2", "交付申請", "Application type dropdown (D2:F2 merged)")

    print(f"\n  [INFO] Other cells contain formulas referencing cost detail sheet")
    print(f"  [INFO] Expected total cost: {TOTAL_COST:,}")
    print(f"  [INFO] Expected target cost: {TARGET_COST:,}")


def fill_cost_detail_sheet(wb):
    """Fill 共通２_太陽光発電 - Cost detail transcription.

    Cell mapping (verified from template openpyxl analysis):
      G3  = application type dropdown (G3:I3 merged)
      D7  = "太陽光発電" (pre-filled, DO NOT WRITE)
      E7  = formula referencing solar sheet G6 - SKIP
      F7  = subsidy rate (empty, WRITE "1/2")
      G7  = formula (VLOOKUP unit cap) - SKIP
      D9:F11 = merged label cell (DO NOT WRITE)

      Cost detail rows 13-37 (continuous, no gap):
        B = cost category (費目区分: 設備費/工事費)
        C = item name (品目名)
        D = detail number (明細番号)
        E = subsidy classification (助成区分: 発電設備)
        F = estimate amount (見積金額)
        G,H = empty (national subsidy columns, not applicable)
        I = formula (都助成対象経費) - SKIP ALL

      Row 38: subtotals (formulas)
      Row 39: 諸経費 (formulas)
      Row 40: 合計 (formulas)
    """
    _, ws = find_sheet_by_name(wb, ["共通２_太陽光発電", "共通2_太陽光発電"])
    print(f"\n{'='*60}")
    print(f"Sheet: {ws.title}")
    print(f"{'='*60}")

    # ------- Application type dropdown -------
    print(f"\n--- (1) Application type dropdown ---")
    safe_write(ws, "G3", "交付申請", "Application type dropdown (G3:I3 merged)")

    # ------- Header data (row 7) -------
    print(f"\n--- (2) Header data ---")
    safe_write(ws, "C7", "なし", "National subsidy co-use dropdown (国等補助金の併用)")
    print(f"  [INFO] D7='太陽光発電' (pre-filled)")
    print(f"  [INFO] E7=formula (system capacity from solar sheet) - SKIP")
    safe_write(ws, "F7", "1/2", "Subsidy rate")
    print(f"  [INFO] G7=formula (unit cap VLOOKUP) - SKIP")
    # D9:F11 is a merged label cell - DO NOT WRITE

    # ------- Equipment cost details -------
    print(f"\n--- (3) Equipment cost details ({len(EQUIPMENT_ITEMS)} items) ---")

    for row, detail_no, item_name, subsidy_class, cost_cat, amount, is_non_target in EQUIPMENT_ITEMS:
        tag = " (non-target)" if is_non_target else ""

        # B column: cost category (設備費 or 工事費)
        safe_write(ws, f"B{row}", cost_cat, f"Cost category{tag}")

        # C column: item name
        safe_write(ws, f"C{row}", item_name, f"Item name{tag}")

        # D column: detail number
        safe_write(ws, f"D{row}", detail_no, f"Detail number{tag}")

        # E column: subsidy classification
        safe_write(ws, f"E{row}", subsidy_class, f"Subsidy class{tag}")

        # F column: estimate amount
        safe_write(ws, f"F{row}", amount, f"Estimate amount: {amount:,}{tag}")

        # H column: 国等補助控除無し都助成対象経費
        # Target items: copy F amount to H (same value)
        # Non-target items (A-5 データロガー, A-6 運用費): leave H empty
        if not is_non_target:
            safe_write(ws, f"H{row}", amount, f"都助成対象経費: {amount:,}")
        else:
            print(f"  [SKIP] H{row}: non-target item ({item_name}) - leave empty")

        # I column: ALL formulas - DO NOT WRITE

    # ------- Summary verification -------
    print(f"\n--- (4) Cost summary verification ---")
    print(f"  [INFO] Row 38 contains SUM formulas (subtotals) - auto-calculated")
    print(f"  [INFO] Row 39 contains 諸経費 formulas (10%) - auto-calculated")
    print(f"  [INFO] Row 40 contains 合計 formulas - auto-calculated")
    print(f"  [VERIFY] Total cost (F column): {TOTAL_COST:,}")
    print(f"  [VERIFY] Non-target expenses: {NON_TARGET_COST:,} (A-5 データロガー 82,000 + A-6 運用費 720,000)")
    print(f"  [VERIFY] Items: {len(EQUIPMENT_ITEMS)} rows")


# =============================================================================
# MAIN
# =============================================================================

def main():
    global INPUT_FILE

    print("=" * 60)
    print("Tokyo Subsidy Auto-Fill - Outside-Tokyo Installation (都外設置)")
    print("Case: 名糖運輸 日高物流センター")
    print("=" * 60)
    print(f"\nInput:  {INPUT_FILE}")
    print(f"Output: {OUTPUT_FILE}")

    # Verify input file exists
    if not os.path.exists(INPUT_FILE):
        print(f"\n[ERROR] Input file not found: {INPUT_FILE}")
        print(f"Please place the empty template at the above path.")
        print(f"\nSearching for alternative template files...")
        # Try alternative names
        alternatives = [
            os.path.join(BASE_DIR, "template.xlsx"),
            os.path.join(BASE_DIR, "r7chisan_togai_yoshiki_20250401.xlsx"),
        ]
        for alt in alternatives:
            if os.path.exists(alt):
                print(f"  Found: {alt}")
                INPUT_FILE = alt
                break
        else:
            print(f"  No alternative found. Exiting.")
            sys.exit(1)

    # Load workbook preserving formulas
    print(f"\nLoading workbook (data_only=False to preserve formulas)...")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=False)
    print(f"Loaded successfully. {len(wb.sheetnames)} sheets found.")
    print(f"Sheets: {wb.sheetnames}")

    # Fill all 4 target sheets
    fill_solar_sheet(wb)
    fill_env_value_sheet(wb)
    fill_overview_sheet(wb)
    fill_cost_detail_sheet(wb)

    # Save output
    print(f"\n{'='*60}")
    print(f"Saving to: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print(f"File saved successfully!")

    # Verify file
    file_size = os.path.getsize(OUTPUT_FILE)
    print(f"Output file size: {file_size:,} bytes")

    # Print final summary
    print(f"\n{'='*60}")
    print("SUMMARY")
    print(f"{'='*60}")
    print(f"  Sheets filled: 4")
    print(f"    1. 共通１－２(太陽光)     - Module/PCS specs + 24 months data")
    print(f"    2. 共通１－４(環境価値)   - 西東京 consumption + certificate plan")
    print(f"    3. 共通２_全体            - Dropdown: 交付申請")
    print(f"    4. 共通２_太陽光発電      - 24 cost items + dropdown")
    print(f"  Equipment items: {len(EQUIPMENT_ITEMS)}")
    print(f"  Total cost: {TOTAL_COST:,}")
    print(f"  Non-target: {NON_TARGET_COST:,}")
    print(f"  Target cost: {TARGET_COST:,}")
    print(f"  Annual generation: {ANNUAL_GENERATION:,.2f} kWh")
    print(f"  Annual consumption (日高): {ANNUAL_CONSUMPTION_HIDAKA:,} kWh")
    print(f"  Annual consumption (西東京): {ANNUAL_CONSUMPTION_NISHITOKYO:,} kWh")
    print(f"  Certificate amount: {ENV_CERTIFICATE_AMOUNT:,} kWh")
    print(f"\n{'='*60}")
    print("COMPLETE - Please open the file in Excel to verify formulas.")
    print("NOTE: Formula cells are NOT recalculated by openpyxl.")
    print("      Open in Excel and press Ctrl+Alt+F9 to recalculate all formulas.")
    print("=" * 60)


if __name__ == "__main__":
    main()
