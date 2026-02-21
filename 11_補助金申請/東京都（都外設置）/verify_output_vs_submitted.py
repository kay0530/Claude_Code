"""
Verification Script: Auto-filled Excel vs Submitted Form (BOX text extraction)
"""

import json
import os
import re
import sys

import openpyxl
from openpyxl.cell.cell import MergedCell

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from auto_fill_subsidy_togai import (
    MODULE_MANUFACTURER, MODULE_MODEL, MODULE_OUTPUT_W, MODULE_COUNT,
    PCS_MANUFACTURER, PCS_MODEL, PCS_RATED_OUTPUT_KW, PCS_UNIT_COUNT,
    CONSUMPTION_HIDAKA, GENERATION_SOLARPRO, CONSUMPTION_NISHITOKYO,
    ENV_COMPANY_NAME, ENV_ADDRESS, ENV_ENERGY_TYPE, ENV_SUBSIDY_RATE,
    ENV_CERTIFICATE_TYPE, ENV_SCHEDULE_LINES, EQUIPMENT_ITEMS,
)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "r7chisan_togai_yoshiki_auto_filled.xlsx")
BOX_TEXT_FILE = os.path.join(
    os.path.expanduser("~"), ".claude", "projects",
    "C--Users-------Desktop-Claude-Code-Demo--claude-worktrees-zen-wilbur",
    "99f07732-b77d-4f1a-afe4-602a3d7c0c34", "tool-results",
    "mcp-c68db0bc-00e1-45ae-ad67-e1829d7faad4-get_file_content-1771382252536.txt",
)


def load_box_text(filepath):
    with open(filepath, "r", encoding="utf-8-sig") as f:
        data = json.load(f)
    if isinstance(data, list) and len(data) > 0:
        return data[0].get("text", "")
    return ""


def get_cell_value(ws, cell_ref):
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merge_range in ws.merged_cells.ranges:
            if cell.coordinate in merge_range:
                top_left = str(merge_range).split(":")[0]
                return ws[top_left].value
    return cell.value


def find_sheet(wb, keywords):
    for name in wb.sheetnames:
        for kw in keywords:
            if kw in name:
                return wb[name]
    return None


def normalize_for_search(value):
    variants = []
    if value is None:
        return variants
    if isinstance(value, (int, float)):
        if isinstance(value, int) or (isinstance(value, float) and value == int(value)):
            int_val = int(value)
            variants.append((f"{int_val:,}", "comma-formatted"))
            variants.append((str(int_val), "plain"))
        else:
            variants.append((f"{value:,.2f}", "comma-formatted 2dp"))
            variants.append((str(value), "plain"))
            int_part = int(value)
            variants.append((f"{int_part:,}", "comma-formatted int part"))
    elif isinstance(value, str):
        variants.append((value, "exact"))
        normalized = re.sub(r"\s+", " ", value).strip()
        if normalized != value:
            variants.append((normalized, "normalized whitespace"))
    else:
        variants.append((str(value), "str cast"))
    return variants


def check_value_in_text(value, box_text, description=""):
    variants = normalize_for_search(value)
    if not variants:
        return False, None, []
    for variant_str, variant_desc in variants:
        if variant_str in box_text:
            return True, f"{variant_str} ({variant_desc})", [v[0] for v in variants]
    return False, None, [v[0] for v in variants]


def build_check_list():
    """Build checks for all values written by auto_fill_subsidy_togai.py."""
    checks = []
    solar_kw = ["共通１－２(太陽光", "共途1-2(太陽光"]
    checks.append((solar_kw, "G12", MODULE_MANUFACTURER, "Module manufacturer"))
    checks.append((solar_kw, "G13", MODULE_MODEL, "Module model"))
    checks.append((solar_kw, "G15", MODULE_OUTPUT_W, "Module output (W)"))
    checks.append((solar_kw, "G16", MODULE_COUNT, "Module count"))
    checks.append((solar_kw, "G39", PCS_MANUFACTURER, "PCS manufacturer"))
    checks.append((solar_kw, "G40", PCS_MODEL, "PCS model"))
    checks.append((solar_kw, "G42", PCS_RATED_OUTPUT_KW, "PCS rated output (kW)"))
    checks.append((solar_kw, "G43", PCS_UNIT_COUNT, "PCS unit count"))

    months_apr_sep = ["apr", "may", "jun", "jul", "aug", "sep"]
    months_oct_mar = ["oct", "nov", "dec", "jan", "feb", "mar"]
    cols = ["E", "F", "G", "H", "I", "J"]

    for col, month in zip(cols, months_apr_sep):
        checks.append((solar_kw, f"{col}126", CONSUMPTION_HIDAKA[month],
                        f"Hidaka consumption {month.capitalize()}"))
    for col, month in zip(cols, months_oct_mar):
        checks.append((solar_kw, f"{col}130", CONSUMPTION_HIDAKA[month],
                        f"Hidaka consumption {month.capitalize()}"))
    for col, month in zip(cols, months_apr_sep):
        checks.append((solar_kw, f"{col}127", GENERATION_SOLARPRO[month],
                        f"SolarPro generation {month.capitalize()}"))
    for col, month in zip(cols, months_oct_mar):
        checks.append((solar_kw, f"{col}131", GENERATION_SOLARPRO[month],
                        f"SolarPro generation {month.capitalize()}"))

    env_kw = ["共通１－４", "共途1-4", "環境価値"]
    checks.append((env_kw, "E3", ENV_COMPANY_NAME, "Company name"))
    checks.append((env_kw, "I3", ENV_ENERGY_TYPE, "Energy type"))
    checks.append((env_kw, "E4", ENV_ADDRESS, "Address"))

    for col, month in zip(cols, months_apr_sep):
        checks.append((env_kw, f"{col}6", CONSUMPTION_NISHITOKYO[month],
                        f"NishiTokyo consumption {month.capitalize()}"))
    for col, month in zip(cols, months_oct_mar):
        checks.append((env_kw, f"{col}8", CONSUMPTION_NISHITOKYO[month],
                        f"NishiTokyo consumption {month.capitalize()}"))

    checks.append((env_kw, "H97", ENV_SUBSIDY_RATE, "Subsidy rate"))
    checks.append((env_kw, "H102", ENV_CERTIFICATE_TYPE, "Certificate type"))

    for row_num, line_text in ENV_SCHEDULE_LINES:
        stripped = line_text.strip()
        if stripped:
            checks.append((env_kw, f"D105(line:{row_num})", stripped,
                            f"Schedule line (row {row_num})"))

    overview_kw = ["共通２_全体", "共途2_全体"]
    checks.append((overview_kw, "D2", "交付申請", "Application type (overview)"))

    cost_kw = ["共通２_太陽光発電", "共途2_太陽光発電"]
    checks.append((cost_kw, "G3", "交付申請", "Application type (cost detail)"))
    checks.append((cost_kw, "F7", "1/2", "Subsidy rate (cost detail)"))

    for row, detail_no, item_name, subsidy_class, cost_cat, amount, is_non_target in EQUIPMENT_ITEMS:
        tag = " [non-target]" if is_non_target else ""
        checks.append((cost_kw, f"B{row}", cost_cat,
                        f"Cost category row {row}{tag}"))
        checks.append((cost_kw, f"C{row}", item_name,
                        f"Item name row {row}{tag}"))
        checks.append((cost_kw, f"D{row}", detail_no,
                        f"Detail number row {row}{tag}"))
        checks.append((cost_kw, f"E{row}", subsidy_class,
                        f"Subsidy class row {row}{tag}"))
        checks.append((cost_kw, f"F{row}", amount,
                        f"Amount row {row}: {amount:,}{tag}"))

    return checks


def main():
    print("=" * 80)
    print("VERIFICATION: Auto-filled Excel vs Submitted Form (BOX text extraction)")
    print("=" * 80)
    print()
    print(f"[1] Loading auto-filled Excel: {EXCEL_FILE}")
    if not os.path.exists(EXCEL_FILE):
        print("    ERROR: File not found!")
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    print(f"    Loaded. Sheets: {wb.sheetnames}")
    print()
    print(f"[2] Loading BOX text extraction: {os.path.basename(BOX_TEXT_FILE)}")
    if not os.path.exists(BOX_TEXT_FILE):
        print("    ERROR: File not found!")
        sys.exit(1)
    box_text = load_box_text(BOX_TEXT_FILE)
    print(f"    Loaded. Text length: {len(box_text):,} characters")
    print()
    checks = build_check_list()
    print(f"[3] Total data cells to verify: {len(checks)}")
    print()
    print("[4] Running verification...")
    print()
    print("-" * 80)

    results = []
    for sheet_kw, cell_ref, expected_value, description in checks:
        ws = find_sheet(wb, sheet_kw)
        sheet_name = ws.title if ws else "NOT FOUND"

        if "D105(line:" in cell_ref:
            excel_val = get_cell_value(ws, "D105") if ws else None
        else:
            excel_val = get_cell_value(ws, cell_ref) if ws else None

        excel_match = False
        if excel_val is not None:
            if isinstance(expected_value, (int, float)):
                try:
                    excel_match = (float(excel_val) == float(expected_value))
                except (ValueError, TypeError):
                    excel_match = (str(excel_val) == str(expected_value))
            elif isinstance(expected_value, str):
                if "D105(line:" in cell_ref:
                    excel_match = (expected_value in str(excel_val))
                else:
                    excel_match = (str(excel_val).strip() == expected_value.strip())
            else:
                excel_match = (excel_val == expected_value)

        box_match, match_variant, tried_variants = check_value_in_text(
            expected_value, box_text, description)

        results.append({
            "sheet": sheet_name, "cell": cell_ref, "description": description,
            "expected": expected_value, "excel_value": excel_val,
            "excel_match": excel_match, "box_match": box_match,
            "match_variant": match_variant, "tried_variants": tried_variants,
        })

    # Print detailed results
    current_sheet = None
    for r in results:
        sheet = r["sheet"]
        if sheet != current_sheet:
            current_sheet = sheet
            print()
            print(f"  === Sheet: {sheet} ===")
            print()
        excel_icon = "OK" if r["excel_match"] else "NG"
        box_icon = "OK" if r["box_match"] else "NG"
        status = f"[Excel:{excel_icon}] [BOX:{box_icon}]"
        cell = r["cell"]
        desc = r["description"]
        expected_str = repr(r["expected"])
        if len(expected_str) > 60:
            expected_str = expected_str[:57] + "..."
        print(f"  {cell:>16s}  {status}  {desc}")
        if not r["excel_match"]:
            print(f"                   >>> Excel expected: {expected_str}")
            excel_str = repr(r["excel_value"])
            if len(excel_str) > 60:
                excel_str = excel_str[:57] + "..."
            print(f"                   >>> Excel actual:   {excel_str}")
        if not r["box_match"]:
            tv = r["tried_variants"]
            print(f"                   >>> BOX miss: tried {tv}")

    # --- Summary ---
    print()
    print("=" * 80)
    print("SUMMARY")
    print("=" * 80)

    total = len(results)
    excel_matches = sum(1 for r in results if r["excel_match"])
    excel_misses = sum(1 for r in results if not r["excel_match"])
    box_matches = sum(1 for r in results if r["box_match"])
    box_misses = sum(1 for r in results if not r["box_match"])
    both_match = sum(1 for r in results if r["excel_match"] and r["box_match"])

    print()
    print(f"  Total data cells checked:     {total}")
    print()
    print("  --- Excel Verification ---")
    print(f"  Matches (Excel == expected):  {excel_matches}")
    print(f"  Misses:                       {excel_misses}")
    print(f"  Accuracy:                     {excel_matches/total*100:.1f}%")
    print()
    print("  --- BOX Text Verification ---")
    print(f"  Matches (value in BOX text):  {box_matches}")
    print(f"  Misses:                       {box_misses}")
    print(f"  Accuracy:                     {box_matches/total*100:.1f}%")
    print()
    print("  --- Combined ---")
    print(f"  Both Excel AND BOX match:     {both_match}/{total} ({both_match/total*100:.1f}%)")

    excel_miss_list = [r for r in results if not r["excel_match"]]
    box_miss_list = [r for r in results if not r["box_match"]]

    if excel_miss_list:
        print()
        print(f"  --- Excel Misses ({len(excel_miss_list)}) ---")
        for r in excel_miss_list:
            es = repr(r["expected"])
            if len(es) > 50:
                es = es[:47] + "..."
            cell = r["cell"]
            sheet = r["sheet"]
            desc = r["description"]
            print(f"    {sheet} / {cell}: {desc}")
            print(f"      Expected: {es}")
            av = repr(r["excel_value"])
            if len(av) > 50:
                av = av[:47] + "..."
            print(f"      Actual:   {av}")

    if box_miss_list:
        print()
        print(f"  --- BOX Misses ({len(box_miss_list)}) ---")
        for r in box_miss_list:
            es = repr(r["expected"])
            if len(es) > 50:
                es = es[:47] + "..."
            cell = r["cell"]
            sheet = r["sheet"]
            desc = r["description"]
            print(f"    {sheet} / {cell}: {desc}")
            print(f"      Expected value: {es}")
            tv = r["tried_variants"]
            print(f"      Searched for:   {tv}")

    if not excel_miss_list and not box_miss_list:
        print()
        print(f"  ALL {total} DATA CELLS VERIFIED SUCCESSFULLY!")
        print("  Every value written by auto_fill_subsidy_togai.py exists in")
        print("  both the output Excel file and the submitted form on BOX.")

    print()
    print("=" * 80)
    print("VERIFICATION COMPLETE")
    print("=" * 80)

    if box_miss_list or excel_miss_list:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
