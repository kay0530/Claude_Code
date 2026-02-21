"""
Verification script for auto-filled Excel file.
Reads and prints values from all 4 target sheets.
"""
import openpyxl
import sys

EXCEL_PATH = r"C:\Users\田中　圭亮\Desktop\Claude_Code_Demo\11_補助金申請\東京都（都外設置）\r7chisan_togai_yoshiki_auto_filled.xlsx"

def find_sheet(wb, keywords_include, keywords_exclude=None):
    """Find a sheet whose name contains all include keywords and none of the exclude keywords."""
    for name in wb.sheetnames:
        if all(k in name for k in keywords_include):
            if keywords_exclude and any(k in name for k in keywords_exclude):
                continue
            return wb[name]
    return None

def cell_val(ws, cell_ref):
    """Return the value of a cell, showing formulas if present."""
    val = ws[cell_ref].value
    return val

def print_range(ws, row, start_col, end_col):
    """Print values in a row from start_col to end_col (letter-based)."""
    from openpyxl.utils import column_index_from_string
    s = column_index_from_string(start_col)
    e = column_index_from_string(end_col)
    vals = []
    for c in range(s, e + 1):
        v = ws.cell(row=row, column=c).value
        vals.append(v)
    return vals

def main():
    print(f"Opening: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    print(f"All sheet names: {wb.sheetnames}\n")

    # =====================================================
    # Sheet 1: 共通１－２(太陽光)
    # Contains 共通 and 太陽光, but NOT 共通２_太陽光発電
    # =====================================================
    print("=" * 70)
    print("Sheet: 共通１－２(太陽光)")
    print("=" * 70)
    ws1 = None
    for name in wb.sheetnames:
        if "共通" in name and "太陽光" in name:
            # Exclude 共通２_太陽光発電
            if "共通２" in name and "太陽光発電" in name:
                continue
            # Exclude 共通１－４ (環境価値)
            if "環境価値" in name or "１－４" in name:
                continue
            ws1 = wb[name]
            print(f"  Found sheet: '{name}'")
            break
    
    if ws1:
        print("\n  --- Module data (G12, G13, G15, G16) ---")
        for r in [12, 13, 15, 16]:
            print(f"    G{r}: {cell_val(ws1, f'G{r}')}")

        print("\n  --- PCS data (G39, G40, G42, G43) ---")
        for r in [39, 40, 42, 43]:
            print(f"    G{r}: {cell_val(ws1, f'G{r}')}")

        print("\n  --- Consumption/Generation rows 126-127 (E-J = Apr-Sep) ---")
        for r in [126, 127]:
            vals = print_range(ws1, r, 'E', 'J')
            print(f"    Row {r} (E-J): {vals}")

        print("\n  --- Consumption/Generation rows 130-131 (E-J = Oct-Mar) ---")
        for r in [130, 131]:
            vals = print_range(ws1, r, 'E', 'J')
            print(f"    Row {r} (E-J): {vals}")
    else:
        print("  *** Sheet NOT FOUND ***")

    # =====================================================
    # Sheet 2: 共通１－４(環境価値)
    # =====================================================
    print("\n" + "=" * 70)
    print("Sheet: 共通１－４(環境価値)")
    print("=" * 70)
    ws2 = None
    for name in wb.sheetnames:
        if "環境価値" in name or "１－４" in name:
            ws2 = wb[name]
            print(f"  Found sheet: '{name}'")
            break

    if ws2:
        print("\n  --- Header info (E3, I3, E4) ---")
        print(f"    E3: {cell_val(ws2, 'E3')}")
        print(f"    I3: {cell_val(ws2, 'I3')}")
        print(f"    E4: {cell_val(ws2, 'E4')}")

        print("\n  --- Monthly consumption row 6 (E-J) ---")
        vals = print_range(ws2, 6, 'E', 'J')
        print(f"    Row 6 (E-J): {vals}")

        print("\n  --- Monthly consumption row 8 (E-J) ---")
        vals = print_range(ws2, 8, 'E', 'J')
        print(f"    Row 8 (E-J): {vals}")

        print("\n  --- Subsidy info (H97, H102) ---")
        print(f"    H97:  {cell_val(ws2, 'H97')}")
        print(f"    H102: {cell_val(ws2, 'H102')}")

        print("\n  --- Schedule text D105 (first 100 chars) ---")
        d105 = cell_val(ws2, 'D105')
        if d105:
            print(f"    D105: {str(d105)[:100]}")
        else:
            print(f"    D105: {d105}")
    else:
        print("  *** Sheet NOT FOUND ***")

    # =====================================================
    # Sheet 3: 共通２_全体
    # =====================================================
    print("\n" + "=" * 70)
    print("Sheet: 共通２_全体")
    print("=" * 70)
    ws3 = None
    for name in wb.sheetnames:
        if "共通２_全体" in name or "共通２_全体" in name:
            ws3 = wb[name]
            print(f"  Found sheet: '{name}'")
            break

    if ws3:
        print("\n  --- D2 ---")
        print(f"    D2: {cell_val(ws3, 'D2')}")
    else:
        print("  *** Sheet NOT FOUND ***")

    # =====================================================
    # Sheet 4: 共通２_太陽光発電
    # =====================================================
    print("\n" + "=" * 70)
    print("Sheet: 共通２_太陽光発電")
    print("=" * 70)
    ws4 = None
    for name in wb.sheetnames:
        if "共通２" in name and "太陽光発電" in name:
            ws4 = wb[name]
            print(f"  Found sheet: '{name}'")
            break

    if ws4:
        print("\n  --- G3, F7 ---")
        print(f"    G3: {cell_val(ws4, 'G3')}")
        print(f"    F7: {cell_val(ws4, 'F7')}")

        print("\n  --- Sample rows (B-F) ---")
        for r in [13, 14, 25, 36]:
            vals = print_range(ws4, r, 'B', 'F')
            print(f"    Row {r} (B-F): {vals}")

        print("\n  --- I13 (should be formula, not overwritten) ---")
        print(f"    I13: {cell_val(ws4, 'I13')}")
    else:
        print("  *** Sheet NOT FOUND ***")

    print("\n" + "=" * 70)
    print("Verification complete.")
    print("=" * 70)

if __name__ == "__main__":
    main()
